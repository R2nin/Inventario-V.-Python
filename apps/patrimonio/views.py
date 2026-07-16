"""
============================================================
VIEWS - Funções que respondem às requisições do usuário
============================================================
Cada função aqui é chamada quando o usuário acessa uma URL.
O fluxo é sempre:
  1. Usuário acessa uma URL no navegador
  2. Django chama a view correspondente
  3. A view processa os dados (lê/salva no banco)
  4. A view retorna um HTML renderizado (template)

Decoradores usados:
  @login_required     → só usuários logados podem acessar
  @staff_member_required → só administradores (is_staff=True)

render(request, 'template.html', contexto)
  → renderiza o template HTML com as variáveis do contexto
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator

from .models import PatrimonioItem, Fornecedor, Localizacao, LogAuditoria, XLSReferenciaItem, ManutencaoSolicitacao, PermissaoConferencia
from .forms import (
    LoginForm, PatrimonioItemForm, FornecedorForm,
    LocalizacaoForm, UsuarioForm, ImportacaoForm, BuscaPatrimonioForm
)
from .utils import processar_arquivo, gerar_pdf_relatorio


# ============================================================
# FUNÇÃO AUXILIAR
# ============================================================
def is_admin(user):
    """Verifica se o usuário é administrador (staff)."""
    return user.is_staff


def pode_conferir_setor(user, local_nome):
    """Verifica se o usuário pode conferir um setor específico (admin ou conferente designado)."""
    if user.is_staff:
        return True
    return PermissaoConferencia.objects.filter(
        usuario=user,
        localizacao__nome=local_nome
    ).exists()


def get_setores_do_conferente(user):
    """Retorna os nomes dos setores que o usuário tem permissão para conferir."""
    if user.is_staff:
        return None  # Admin vê todos
    return list(
        PermissaoConferencia.objects.filter(usuario=user)
        .values_list('localizacao__nome', flat=True)
    )


# ============================================================
# AUTENTICAÇÃO
# ============================================================

def login_view(request):
    """
    Exibe e processa o formulário de login.

    GET  → exibe o formulário vazio
    POST → valida as credenciais e faz o login
    """
    # Se já está logado, redireciona para o dashboard
    if request.user.is_authenticated:
        return redirect('dashboard')

    form = LoginForm()

    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # authenticate() verifica as credenciais no banco
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)  # Cria a sessão do usuário

                # Registra o login no log de auditoria
                LogAuditoria.registrar(
                    acao=LogAuditoria.ACAO_LOGIN,
                    tipo_entidade=LogAuditoria.ENTIDADE_SISTEMA,
                    descricao=f'Usuário "{user.username}" fez login no sistema.',
                    usuario=user,
                )

                messages.success(request, f'Bem-vindo, {user.get_full_name() or user.username}!')
                # Redireciona para a URL que o usuário tentou acessar, ou para o dashboard
                proximo = request.GET.get('next', 'dashboard')
                return redirect(proximo)
            else:
                messages.error(request, 'Usuário ou senha inválidos. Tente novamente.')

    return render(request, 'patrimonio/login.html', {'form': form})


@login_required
def logout_view(request):
    """Faz o logout do usuário e redireciona para o login."""
    LogAuditoria.registrar(
        acao=LogAuditoria.ACAO_LOGOUT,
        tipo_entidade=LogAuditoria.ENTIDADE_SISTEMA,
        descricao=f'Usuário "{request.user.username}" saiu do sistema.',
        usuario=request.user,
    )
    logout(request)
    messages.info(request, 'Você saiu do sistema.')
    return redirect('login')


# ============================================================
# DASHBOARD
# ============================================================

@login_required
def dashboard(request):
    """
    Página principal com estatísticas gerais do patrimônio.
    Acessível para todos os usuários logados.
    """
    # Calcula as estatísticas
    total_itens      = PatrimonioItem.objects.count()
    itens_ativos     = PatrimonioItem.objects.filter(status='ativo').count()
    itens_manutencao = PatrimonioItem.objects.filter(status='manutencao').count()
    itens_baixados   = PatrimonioItem.objects.filter(status='baixado').count()

    # Soma o valor total de todos os itens
    valor_total = PatrimonioItem.objects.aggregate(
        total=Sum('valor')
    )['total'] or 0

    # Contagem por localização (para o gráfico)
    por_localizacao = (
        PatrimonioItem.objects
        .values('localizacao__nome')
        .annotate(quantidade=Count('id'))
        .order_by('-quantidade')[:10]  # Top 10 localizações
    )

    # Últimos 5 itens adicionados
    ultimos_itens = PatrimonioItem.objects.select_related(
        'localizacao', 'fornecedor'
    ).order_by('-criado_em')[:5]

    # Últimas 10 ações no log
    ultimos_logs = LogAuditoria.objects.order_by('-criado_em')[:10]

    # Movimentações feitas por conferentes (não-admins) — notificação para admin
    transferencias_conferente = 0
    if request.user.is_staff:
        transferencias_conferente = LogAuditoria.objects.filter(
            acao=LogAuditoria.ACAO_TRANSFERIR,
            descricao__startswith='[CONFERENTE]',
        ).count()

    contexto = {
        'total_itens': total_itens,
        'itens_ativos': itens_ativos,
        'itens_manutencao': itens_manutencao,
        'itens_baixados': itens_baixados,
        'valor_total': valor_total,
        'por_localizacao': por_localizacao,
        'ultimos_itens': ultimos_itens,
        'ultimos_logs': ultimos_logs,
        'transferencias_conferente': transferencias_conferente,
        'pagina_ativa': 'dashboard',
    }
    return render(request, 'patrimonio/dashboard.html', contexto)


# ============================================================
# PATRIMÔNIO - CRUD
# ============================================================

@login_required
def patrimonio_lista(request):
    """
    Lista todos os itens patrimoniais com busca e filtros.
    Suporta paginação (20 itens por página).
    """
    form_busca = BuscaPatrimonioForm(request.GET)
    itens = PatrimonioItem.objects.select_related('localizacao', 'fornecedor').all()

    # Aplica filtros se o formulário for válido
    if form_busca.is_valid():
        busca = form_busca.cleaned_data.get('busca')
        status = form_busca.cleaned_data.get('status')
        localizacao = form_busca.cleaned_data.get('localizacao')

        if busca:
            # Q() permite combinar condições com OR (|)
            itens = itens.filter(
                Q(nome__icontains=busca) |
                Q(numero_chapa__icontains=busca) |
                Q(responsavel__icontains=busca) |
                Q(categoria__icontains=busca)
            )
        if status:
            itens = itens.filter(status=status)
        if localizacao:
            itens = itens.filter(localizacao=localizacao)

    # Ordenação
    ordem = request.GET.get('ordem', 'numero_chapa')
    direcao = request.GET.get('dir', 'asc')
    if direcao == 'desc':
        ordem = f'-{ordem}'
    itens = itens.order_by(ordem)

    # Paginação: 20 itens por página
    paginador = Paginator(itens, 20)
    pagina_num = request.GET.get('pagina', 1)
    pagina = paginador.get_page(pagina_num)

    contexto = {
        'itens': pagina,
        'form_busca': form_busca,
        'total': itens.count() if hasattr(itens, 'count') else len(itens),
        'pagina_ativa': 'patrimonio',
    }
    return render(request, 'patrimonio/patrimonio_lista.html', contexto)


@login_required
def patrimonio_detalhe(request, pk):
    """Exibe os detalhes completos de um item patrimonial."""
    # get_object_or_404 retorna 404 automaticamente se não encontrar
    item = get_object_or_404(PatrimonioItem.objects.select_related(
        'localizacao', 'fornecedor'
    ), pk=pk)

    # Histórico de logs deste item
    logs = LogAuditoria.objects.filter(
        entidade_id=str(pk),
        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO
    ).order_by('-criado_em')[:20]

    return render(request, 'patrimonio/patrimonio_detalhe.html', {
        'item': item,
        'logs': logs,
        'pagina_ativa': 'patrimonio',
    })


@login_required
def patrimonio_criar(request):
    """
    Cria um novo item patrimonial.
    Suporta parâmetros GET opcionais para o modo de conferência:
      ?local=NomeDaSala  → trava a localização pré-selecionada
      ?chapa=1234        → preenche o número de chapa e dispara auto-fill do XLS
    Ao salvar no modo conferência, volta para a tela da sala.

    Acesso:
      - Admin: acesso total (criar qualquer item em qualquer local)
      - Conferente: acesso restrito ao contexto da conferência (?local= obrigatório
        e deve ser um setor autorizado para o usuário)
    """
    proximo_chapa = PatrimonioItem.proximo_numero_chapa()

    # Modo conferência: localização travada passada via GET (ou via POST no safety net)
    local_nome    = request.GET.get('local', '').strip()
    chapa_inicial = request.GET.get('chapa', '')
    local_travado = None
    if local_nome:
        local_travado = Localizacao.objects.filter(nome=local_nome).first()

    # Verificação de acesso: admin tem acesso total; conferente só pode criar
    # itens no contexto de um setor que tem permissão para conferir.
    if not request.user.is_staff:
        if not local_nome or not pode_conferir_setor(request.user, local_nome):
            messages.error(request, 'Sem permissão para cadastrar itens fora do contexto da conferência do seu setor.')
            return redirect('conferencia_inicio')

    if request.method == 'POST':
        # Safety net para modo conferência: se a chapa já existe, transfere em vez de criar
        voltar_local = request.POST.get('voltar_local', '').strip()
        chapa_str = request.POST.get('numero_chapa', '')
        if voltar_local and chapa_str:
            try:
                chapa = int(chapa_str)
                existente = PatrimonioItem.objects.filter(numero_chapa=chapa).first()
                if existente:
                    from django.urls import reverse
                    loc, _ = Localizacao.objects.get_or_create(nome=voltar_local)
                    loc_anterior = existente.localizacao.nome if existente.localizacao else 'sem localização'
                    existente.localizacao = loc
                    existente.save(update_fields=['localizacao', 'atualizado_em'])
                    LogAuditoria.registrar(
                        acao=LogAuditoria.ACAO_EDITAR,
                        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
                        descricao=f'Conferência: transferiu [{existente.numero_chapa}] {existente.nome} de "{loc_anterior}" para "{voltar_local}".',
                        usuario=request.user,
                        entidade_id=str(existente.pk),
                        entidade_nome=existente.nome,
                    )
                    messages.success(request, f'Item [{chapa}] já cadastrado — localização atualizada para "{voltar_local}".')
                    return redirect(f"{reverse('conferencia_sala')}?local={voltar_local}")
            except (ValueError, TypeError):
                pass

        form = PatrimonioItemForm(request.POST)
        if form.is_valid():
            try:
                quantidade = max(1, min(500, int(request.POST.get('quantidade', 1))))
            except (ValueError, TypeError):
                quantidade = 1

            if quantidade == 1:
                # Criação individual (fluxo original)
                item = form.save(commit=False)
                item.criado_por = request.user
                item.criado_por_nome = request.user.get_full_name() or request.user.username
                item.save()
                LogAuditoria.registrar(
                    acao=LogAuditoria.ACAO_CRIAR,
                    tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
                    descricao=f'Criou o item patrimonial "{item.nome}" (chapa {item.numero_chapa}).',
                    usuario=request.user,
                    entidade_id=item.pk,
                    entidade_nome=item.nome,
                )
                messages.success(request, f'Item "{item.nome}" criado com sucesso!')

                # Se veio da conferência, grava chapa na sessão e volta para a sala
                if voltar_local:
                    request.session['ultima_chapa_conferencia'] = item.numero_chapa
                    from django.urls import reverse
                    return redirect(f"{reverse('conferencia_sala')}?local={voltar_local}#chapa-{item.numero_chapa}")
                return redirect('patrimonio_detalhe', pk=item.pk)

            else:
                # Criação em lote: mesmo nome/local/data/descrição, chapas sequenciais
                chapa_base  = form.cleaned_data['numero_chapa']
                nome        = form.cleaned_data['nome']
                localizacao = form.cleaned_data.get('localizacao')
                data_aq     = form.cleaned_data.get('data_aquisicao')
                descricao   = form.cleaned_data.get('descricao', '')
                usuario_nome = request.user.get_full_name() or request.user.username

                criados    = []
                duplicados = []
                for i in range(quantidade):
                    chapa_atual = chapa_base + i
                    if PatrimonioItem.objects.filter(numero_chapa=chapa_atual).exists():
                        duplicados.append(chapa_atual)
                        continue
                    novo = PatrimonioItem(
                        numero_chapa=chapa_atual,
                        nome=nome,
                        localizacao=localizacao,
                        data_aquisicao=data_aq,
                        descricao=descricao,
                        criado_por=request.user,
                        criado_por_nome=usuario_nome,
                    )
                    novo.save()
                    LogAuditoria.registrar(
                        acao=LogAuditoria.ACAO_CRIAR,
                        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
                        descricao=f'Criou em lote: "{nome}" (chapa {chapa_atual}) — lote de {quantidade}.',
                        usuario=request.user,
                        entidade_id=novo.pk,
                        entidade_nome=novo.nome,
                    )
                    criados.append(chapa_atual)

                if criados:
                    messages.success(
                        request,
                        f'{len(criados)} item(ns) criado(s) com sucesso '
                        f'(chapas {criados[0]} a {criados[-1]}).'
                    )
                if duplicados:
                    messages.warning(
                        request,
                        f'Chapas já existentes (ignoradas): {", ".join(map(str, duplicados))}.'
                    )
                return redirect('patrimonio_lista')
        else:
            messages.error(request, 'Corrija os erros abaixo antes de salvar.')
    else:
        initial = {'numero_chapa': chapa_inicial or proximo_chapa}
        if local_travado:
            initial['localizacao'] = local_travado
        form = PatrimonioItemForm(initial=initial)

    return render(request, 'patrimonio/patrimonio_form.html', {
        'form':          form,
        'titulo':        'Novo Item Patrimonial',
        'pagina_ativa':  'patrimonio',
        'local_travado': local_travado,
        'chapa_inicial': chapa_inicial,
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def patrimonio_editar(request, pk):
    """
    Edita um item patrimonial existente.
    Apenas administradores podem editar.
    """
    item = get_object_or_404(PatrimonioItem, pk=pk)

    if request.method == 'POST':
        form = PatrimonioItemForm(request.POST, instance=item)
        if form.is_valid():
            item_atualizado = form.save()

            LogAuditoria.registrar(
                acao=LogAuditoria.ACAO_EDITAR,
                tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
                descricao=f'Editou o item patrimonial "{item_atualizado.nome}" (chapa {item_atualizado.numero_chapa}).',
                usuario=request.user,
                entidade_id=item.pk,
                entidade_nome=item.nome,
            )

            messages.success(request, f'Item "{item.nome}" atualizado com sucesso!')
            return redirect('patrimonio_detalhe', pk=item.pk)
        else:
            messages.error(request, 'Corrija os erros abaixo antes de salvar.')
    else:
        form = PatrimonioItemForm(instance=item)

    return render(request, 'patrimonio/patrimonio_form.html', {
        'form': form,
        'item': item,
        'titulo': f'Editar: {item.nome}',
        'pagina_ativa': 'patrimonio',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def patrimonio_deletar(request, pk):
    """
    Deleta um item patrimonial.
    Só aceita POST para evitar deleções acidentais por GET.
    """
    item = get_object_or_404(PatrimonioItem, pk=pk)

    if request.method == 'POST':
        nome = item.nome
        chapa = item.numero_chapa
        item.delete()

        LogAuditoria.registrar(
            acao=LogAuditoria.ACAO_DELETAR,
            tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
            descricao=f'Deletou o item patrimonial "{nome}" (chapa {chapa}).',
            usuario=request.user,
            entidade_nome=nome,
        )

        messages.success(request, f'Item "{nome}" deletado com sucesso.')
        return redirect('patrimonio_lista')

    # GET: exibe página de confirmação
    return render(request, 'patrimonio/confirmar_delete.html', {
        'objeto': item,
        'tipo': 'item patrimonial',
        'cancelar_url': 'patrimonio_detalhe',
        'cancelar_pk': pk,
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def patrimonio_importar(request):
    """
    Importa itens patrimoniais a partir de arquivo Excel ou CSV.
    GET  → exibe formulário de upload
    POST → processa o arquivo e salva os itens
    """
    form = ImportacaoForm()
    preview = None
    erros = []

    if request.method == 'POST':
        # Botão "processar" - faz o preview do arquivo
        if 'processar' in request.POST:
            form = ImportacaoForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    import uuid
                    from pathlib import Path
                    from django.conf import settings

                    # Salva o arquivo em disco para reler no confirmar (evita sessão pesada)
                    tmp_dir = Path(settings.MEDIA_ROOT) / 'tmp_import'
                    tmp_dir.mkdir(parents=True, exist_ok=True)
                    ext = Path(request.FILES['arquivo'].name).suffix
                    tmp_path = tmp_dir / f"{uuid.uuid4().hex}{ext}"
                    with open(tmp_path, 'wb') as f:
                        for chunk in request.FILES['arquivo'].chunks():
                            f.write(chunk)

                    with open(tmp_path, 'rb') as f:
                        itens_preview = processar_arquivo(f)

                    request.session['importacao_tmp'] = str(tmp_path)
                    preview = itens_preview
                    messages.info(request, f'{len(itens_preview)} itens encontrados. Revise e confirme a importação.')
                except Exception as e:
                    messages.error(request, f'Erro ao processar arquivo: {e}')

        # Botão "confirmar" ou "substituir" - salva os itens no banco
        elif 'confirmar' in request.POST or 'substituir' in request.POST:
            import datetime
            from decimal import Decimal
            from pathlib import Path

            substituir_tudo = 'substituir' in request.POST

            tmp_path_str = request.session.get('importacao_tmp')
            if not tmp_path_str or not Path(tmp_path_str).exists():
                messages.error(request, 'Arquivo temporário não encontrado. Envie o arquivo novamente.')
                return redirect('patrimonio_importar')

            try:
                with open(tmp_path_str, 'rb') as f:
                    dados = processar_arquivo(f)
            except Exception as e:
                messages.error(request, f'Erro ao reler o arquivo: {e}')
                return redirect('patrimonio_importar')

            erros_import = []

            # Modo "substituir": apaga todos os itens e localizações antes de importar
            if substituir_tudo:
                PatrimonioItem.objects.all().delete()
                Localizacao.objects.all().delete()

            # Pré-carrega chapas existentes (1 query)
            chapas_existentes = set(PatrimonioItem.objects.values_list('numero_chapa', flat=True))

            # Pré-cria localizações únicas (bulk)
            nomes_loc = {
                item.get('localizacao_nome') for item in dados
                if item.get('localizacao_nome')
            }
            existentes = set(Localizacao.objects.values_list('nome', flat=True))
            novas = [Localizacao(nome=n[:255]) for n in nomes_loc if n not in existentes]
            if novas:
                Localizacao.objects.bulk_create(novas, ignore_conflicts=True)
            cache_loc = {loc.nome: loc for loc in Localizacao.objects.all()}

            # Monta objetos sem tocar no banco
            proxima_chapa = PatrimonioItem.proximo_numero_chapa()
            objetos = []
            for item in dados:
                try:
                    chapa = item.get('numero_chapa')
                    if chapa:
                        if chapa in chapas_existentes:
                            erros_import.append(f'Chapa {chapa} já existe - pulado.')
                            continue
                        chapas_existentes.add(chapa)
                    else:
                        chapa = proxima_chapa
                        proxima_chapa += 1

                    loc_nome = item.get('localizacao_nome')
                    localizacao = cache_loc.get(loc_nome) if loc_nome else None

                    data_val = item.get('data_aquisicao')
                    if isinstance(data_val, str):
                        try:
                            data_val = datetime.date.fromisoformat(data_val)
                        except Exception:
                            data_val = None

                    objetos.append(PatrimonioItem(
                        numero_chapa=chapa,
                        nome=(item.get('nome', '') or '')[:255],
                        categoria=(item.get('categoria', '') or '')[:100],
                        localizacao=localizacao,
                        responsavel=(item.get('responsavel', '') or '')[:255],
                        data_aquisicao=data_val,
                        valor=item.get('valor'),
                        status=item.get('status', 'ativo') or 'ativo',
                        descricao=item.get('descricao', '') or '',
                    ))
                except Exception as e:
                    erros_import.append(f'Linha {item.get("_linha", "?")}: {e}')

            # Insere tudo de uma vez
            try:
                criados = PatrimonioItem.objects.bulk_create(objetos, batch_size=500, ignore_conflicts=True)
                salvos = len(criados)
            except Exception as e:
                messages.error(request, f'Erro ao salvar no banco: {e}')
                return redirect('patrimonio_importar')

            # Limpa arquivo temporário e sessão
            try:
                Path(tmp_path_str).unlink(missing_ok=True)
            except Exception:
                pass
            request.session.pop('importacao_tmp', None)

            acao_desc = 'Substituiu' if substituir_tudo else 'Importou'
            LogAuditoria.registrar(
                acao=LogAuditoria.ACAO_IMPORTAR,
                tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
                descricao=f'{acao_desc} {salvos} itens via arquivo. {len(erros_import)} erros.',
                usuario=request.user,
            )

            if salvos:
                messages.success(request, f'{salvos} itens importados com sucesso!')
            if erros_import:
                for e in erros_import[:5]:
                    messages.warning(request, e)

            return redirect('patrimonio_lista')

    return render(request, 'patrimonio/patrimonio_importar.html', {
        'form': form,
        'preview': preview,
        'erros': erros,
        'pagina_ativa': 'patrimonio',
    })


@login_required
def patrimonio_pdf(request):
    """Gera e retorna o PDF do relatório de patrimônio."""
    itens = PatrimonioItem.objects.select_related('localizacao', 'fornecedor').all()

    # Aplica filtros vindos dos parâmetros GET
    filtros = {}
    status = request.GET.get('status')
    localizacao_id = request.GET.get('localizacao')

    if status:
        itens = itens.filter(status=status)
        labels = {'ativo': 'Ativo', 'manutencao': 'Manutenção', 'baixado': 'Baixado'}
        filtros['Status'] = labels.get(status, status)
    if localizacao_id:
        itens = itens.filter(localizacao_id=localizacao_id)
        try:
            filtros['Localização'] = Localizacao.objects.get(pk=localizacao_id).nome
        except Localizacao.DoesNotExist:
            pass

    return gerar_pdf_relatorio(list(itens), filtros=filtros or None)


# ============================================================
# FORNECEDORES
# ============================================================

@login_required
def fornecedor_lista(request):
    """Lista todos os fornecedores com busca."""
    busca = request.GET.get('busca', '')
    fornecedores = Fornecedor.objects.all()

    if busca:
        fornecedores = fornecedores.filter(
            Q(nome__icontains=busca) |
            Q(cnpj__icontains=busca) |
            Q(contato__icontains=busca)
        )

    paginador = Paginator(fornecedores, 20)
    pagina = paginador.get_page(request.GET.get('pagina', 1))

    return render(request, 'patrimonio/fornecedor_lista.html', {
        'fornecedores': pagina,
        'busca': busca,
        'pagina_ativa': 'fornecedores',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def fornecedor_criar(request):
    """Cria um novo fornecedor."""
    if request.method == 'POST':
        form = FornecedorForm(request.POST)
        if form.is_valid():
            fornecedor = form.save()
            LogAuditoria.registrar(
                acao=LogAuditoria.ACAO_CRIAR,
                tipo_entidade=LogAuditoria.ENTIDADE_FORNECEDOR,
                descricao=f'Criou o fornecedor "{fornecedor.nome}".',
                usuario=request.user,
                entidade_id=fornecedor.pk,
                entidade_nome=fornecedor.nome,
            )
            messages.success(request, f'Fornecedor "{fornecedor.nome}" criado com sucesso!')
            return redirect('fornecedor_lista')
        else:
            messages.error(request, 'Corrija os erros abaixo.')
    else:
        form = FornecedorForm()

    return render(request, 'patrimonio/fornecedor_form.html', {
        'form': form,
        'titulo': 'Novo Fornecedor',
        'pagina_ativa': 'fornecedores',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def fornecedor_editar(request, pk):
    """Edita um fornecedor existente."""
    fornecedor = get_object_or_404(Fornecedor, pk=pk)

    if request.method == 'POST':
        form = FornecedorForm(request.POST, instance=fornecedor)
        if form.is_valid():
            form.save()
            LogAuditoria.registrar(
                acao=LogAuditoria.ACAO_EDITAR,
                tipo_entidade=LogAuditoria.ENTIDADE_FORNECEDOR,
                descricao=f'Editou o fornecedor "{fornecedor.nome}".',
                usuario=request.user,
                entidade_id=fornecedor.pk,
                entidade_nome=fornecedor.nome,
            )
            messages.success(request, f'Fornecedor "{fornecedor.nome}" atualizado!')
            return redirect('fornecedor_lista')
        else:
            messages.error(request, 'Corrija os erros abaixo.')
    else:
        form = FornecedorForm(instance=fornecedor)

    return render(request, 'patrimonio/fornecedor_form.html', {
        'form': form,
        'objeto': fornecedor,
        'titulo': f'Editar: {fornecedor.nome}',
        'pagina_ativa': 'fornecedores',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def fornecedor_deletar(request, pk):
    """Deleta um fornecedor (com confirmação)."""
    fornecedor = get_object_or_404(Fornecedor, pk=pk)

    if request.method == 'POST':
        nome = fornecedor.nome
        fornecedor.delete()
        LogAuditoria.registrar(
            acao=LogAuditoria.ACAO_DELETAR,
            tipo_entidade=LogAuditoria.ENTIDADE_FORNECEDOR,
            descricao=f'Deletou o fornecedor "{nome}".',
            usuario=request.user,
            entidade_nome=nome,
        )
        messages.success(request, f'Fornecedor "{nome}" deletado.')
        return redirect('fornecedor_lista')

    return render(request, 'patrimonio/confirmar_delete.html', {
        'objeto': fornecedor,
        'tipo': 'fornecedor',
        'cancelar_url': 'fornecedor_lista',
    })


# ============================================================
# LOCALIZAÇÕES
# ============================================================

@login_required
def localizacao_lista(request):
    """Lista todas as localizações."""
    busca = request.GET.get('busca', '')
    localizacoes = Localizacao.objects.annotate(total_itens=Count('patrimonioitem')).order_by('nome')

    if busca:
        localizacoes = localizacoes.filter(nome__icontains=busca)

    return render(request, 'patrimonio/localizacao_lista.html', {
        'localizacoes': localizacoes,
        'busca': busca,
        'pagina_ativa': 'localizacoes',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def localizacao_criar(request):
    """Cria uma nova localização."""
    if request.method == 'POST':
        form = LocalizacaoForm(request.POST)
        if form.is_valid():
            loc = form.save()
            LogAuditoria.registrar(
                acao=LogAuditoria.ACAO_CRIAR,
                tipo_entidade=LogAuditoria.ENTIDADE_LOCALIZACAO,
                descricao=f'Criou a localização "{loc.nome}".',
                usuario=request.user,
                entidade_id=loc.pk,
                entidade_nome=loc.nome,
            )
            messages.success(request, f'Localização "{loc.nome}" criada!')
            return redirect('localizacao_lista')
    else:
        form = LocalizacaoForm()

    return render(request, 'patrimonio/localizacao_form.html', {
        'form': form,
        'titulo': 'Nova Localização',
        'pagina_ativa': 'localizacoes',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def localizacao_editar(request, pk):
    """Edita uma localização existente."""
    loc = get_object_or_404(Localizacao, pk=pk)

    if request.method == 'POST':
        nome_antigo = loc.nome
        form = LocalizacaoForm(request.POST, instance=loc)
        if form.is_valid():
            form.save()
            nome_novo = loc.nome

            # Se o nome mudou, atualiza os registros do XLS de referência
            if nome_antigo != nome_novo:
                atualizados = XLSReferenciaItem.objects.filter(
                    local=nome_antigo
                ).update(local=nome_novo)
            else:
                atualizados = 0

            LogAuditoria.registrar(
                acao=LogAuditoria.ACAO_EDITAR,
                tipo_entidade=LogAuditoria.ENTIDADE_LOCALIZACAO,
                descricao=f'Editou a localização "{nome_antigo}" → "{nome_novo}".',
                usuario=request.user,
                entidade_id=loc.pk,
                entidade_nome=nome_novo,
            )
            if atualizados:
                messages.success(request, f'Localização renomeada de "{nome_antigo}" para "{nome_novo}". {atualizados} item(ns) do XLS atualizados.')
            else:
                messages.success(request, f'Localização "{nome_novo}" atualizada!')
            return redirect('localizacao_lista')
    else:
        form = LocalizacaoForm(instance=loc)

    return render(request, 'patrimonio/localizacao_form.html', {
        'form': form,
        'objeto': loc,
        'titulo': f'Editar: {loc.nome}',
        'pagina_ativa': 'localizacoes',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def localizacao_deletar(request, pk):
    """Deleta uma localização (com confirmação)."""
    loc = get_object_or_404(Localizacao, pk=pk)

    if request.method == 'POST':
        nome = loc.nome
        loc.delete()
        LogAuditoria.registrar(
            acao=LogAuditoria.ACAO_DELETAR,
            tipo_entidade=LogAuditoria.ENTIDADE_LOCALIZACAO,
            descricao=f'Deletou a localização "{nome}".',
            usuario=request.user,
            entidade_nome=nome,
        )
        messages.success(request, f'Localização "{nome}" deletada.')
        return redirect('localizacao_lista')

    return render(request, 'patrimonio/confirmar_delete.html', {
        'objeto': loc,
        'tipo': 'localização',
        'cancelar_url': 'localizacao_lista',
    })


# ============================================================
# USUÁRIOS (só admin)
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='dashboard')
def usuario_lista(request):
    """Lista todos os usuários do sistema."""
    busca = request.GET.get('busca', '')
    usuarios = User.objects.all().order_by('username')

    if busca:
        usuarios = usuarios.filter(
            Q(username__icontains=busca) |
            Q(first_name__icontains=busca) |
            Q(last_name__icontains=busca) |
            Q(email__icontains=busca)
        )

    return render(request, 'patrimonio/usuario_lista.html', {
        'usuarios': usuarios,
        'busca': busca,
        'pagina_ativa': 'usuarios',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def usuario_criar(request):
    """Cria um novo usuário."""
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            usuario = form.save()
            LogAuditoria.registrar(
                acao=LogAuditoria.ACAO_CRIAR,
                tipo_entidade=LogAuditoria.ENTIDADE_USUARIO,
                descricao=f'Criou o usuário "{usuario.username}".',
                usuario=request.user,
                entidade_id=usuario.pk,
                entidade_nome=usuario.username,
            )
            messages.success(request, f'Usuário "{usuario.username}" criado com sucesso!')
            return redirect('usuario_lista')
        else:
            messages.error(request, 'Corrija os erros abaixo.')
    else:
        form = UsuarioForm()

    return render(request, 'patrimonio/usuario_form.html', {
        'form': form,
        'titulo': 'Novo Usuário',
        'pagina_ativa': 'usuarios',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def usuario_deletar(request, pk):
    """Desativa um usuário (não deleta para preservar logs)."""
    usuario = get_object_or_404(User, pk=pk)

    # Não permite desativar a própria conta
    if usuario == request.user:
        messages.error(request, 'Você não pode desativar sua própria conta.')
        return redirect('usuario_lista')

    if request.method == 'POST':
        usuario.is_active = False  # Desativa ao invés de deletar
        usuario.save()
        LogAuditoria.registrar(
            acao=LogAuditoria.ACAO_DELETAR,
            tipo_entidade=LogAuditoria.ENTIDADE_USUARIO,
            descricao=f'Desativou o usuário "{usuario.username}".',
            usuario=request.user,
            entidade_id=usuario.pk,
            entidade_nome=usuario.username,
        )
        messages.success(request, f'Usuário "{usuario.username}" desativado.')
        return redirect('usuario_lista')

    return render(request, 'patrimonio/confirmar_delete.html', {
        'objeto': usuario,
        'tipo': 'usuário',
        'cancelar_url': 'usuario_lista',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def usuario_permissoes(request, pk):
    """
    Gerencia quais setores um usuário pode conferir.
    GET  → exibe todas as localizações com checkboxes indicando quais já tem acesso.
    POST → salva as novas permissões (remove as desmarcadas, adiciona as marcadas).
    """
    usuario = get_object_or_404(User, pk=pk)

    # Admins já têm acesso total — não faz sentido configurar permissões
    if usuario.is_staff:
        messages.info(request, f'"{usuario.username}" é admin e já tem acesso a todos os setores.')
        return redirect('usuario_lista')

    localizacoes = Localizacao.objects.all().order_by('nome')
    permissoes_atuais = set(
        PermissaoConferencia.objects.filter(usuario=usuario)
        .values_list('localizacao_id', flat=True)
    )

    if request.method == 'POST':
        ids_marcados = set(int(x) for x in request.POST.getlist('localizacoes'))

        # Remover permissões desmarcadas
        PermissaoConferencia.objects.filter(
            usuario=usuario
        ).exclude(localizacao_id__in=ids_marcados).delete()

        # Adicionar novas permissões
        for loc_id in ids_marcados:
            if loc_id not in permissoes_atuais:
                try:
                    loc = Localizacao.objects.get(pk=loc_id)
                    PermissaoConferencia.objects.get_or_create(
                        usuario=usuario,
                        localizacao=loc,
                        defaults={'criado_por': request.user}
                    )
                except Localizacao.DoesNotExist:
                    pass

        LogAuditoria.registrar(
            acao=LogAuditoria.ACAO_EDITAR,
            tipo_entidade=LogAuditoria.ENTIDADE_USUARIO,
            descricao=(
                f'Atualizou permissões de conferência de "{usuario.username}": '
                f'{len(ids_marcados)} setor(es) atribuído(s).'
            ),
            usuario=request.user,
            entidade_id=usuario.pk,
            entidade_nome=usuario.username,
        )
        messages.success(request, f'Permissões de "{usuario.username}" atualizadas.')
        return redirect('usuario_lista')

    return render(request, 'patrimonio/usuario_permissoes.html', {
        'usuario_alvo':     usuario,
        'localizacoes':     localizacoes,
        'permissoes_atuais': permissoes_atuais,
        'pagina_ativa':     'usuarios',
    })


# ============================================================
# LOGS DE AUDITORIA
# ============================================================

@login_required
def log_lista(request):
    """Lista o histórico de auditoria com filtros."""
    busca = request.GET.get('busca', '')
    tipo = request.GET.get('tipo', '')
    acao = request.GET.get('acao', '')

    logs = LogAuditoria.objects.select_related('usuario').all()

    if busca:
        logs = logs.filter(
            Q(descricao__icontains=busca) |
            Q(usuario_nome__icontains=busca) |
            Q(entidade_nome__icontains=busca)
        )
    if tipo:
        logs = logs.filter(tipo_entidade=tipo)
    if acao:
        logs = logs.filter(acao=acao)

    paginador = Paginator(logs, 30)
    pagina = paginador.get_page(request.GET.get('pagina', 1))

    return render(request, 'patrimonio/log_lista.html', {
        'logs': pagina,
        'busca': busca,
        'tipo_filtro': tipo,
        'acao_filtro': acao,
        'tipos': LogAuditoria.ENTIDADE_CHOICES,
        'acoes': LogAuditoria.ACAO_CHOICES,
        'pagina_ativa': 'logs',
    })


# ============================================================
# QR CODE
# ============================================================

@login_required
def scanner_qrcode(request):
    """
    Página com leitor de QR Code via câmera do dispositivo.
    Permite definir uma localização antes de escanear e atualiza
    a localização de cada item confirmado. Os itens lidos ficam
    numa tabela e podem ser exportados como CSV.
    """
    localizacoes = Localizacao.objects.all().order_by('nome')
    return render(request, 'patrimonio/scanner_qrcode.html', {
        'pagina_ativa': 'scanner',
        'localizacoes': localizacoes,
    })


@login_required
def api_item_por_chapa(request, chapa):
    """
    API JSON: retorna dados do item pelo número de chapa.
    Usado pelo scanner para mostrar o painel de confirmação.
    """
    from django.http import JsonResponse
    try:
        item = PatrimonioItem.objects.select_related('localizacao').get(numero_chapa=chapa)
        status_labels = dict(PatrimonioItem.STATUS_CHOICES)
        return JsonResponse({
            'ok': True,
            'pk': item.pk,
            'chapa': item.numero_chapa,
            'nome': item.nome,
            'localizacao_id': item.localizacao_id,
            'localizacao_nome': item.localizacao.nome if item.localizacao else '',
            'status': item.status,
            'status_display': status_labels.get(item.status, item.status),
        })
    except PatrimonioItem.DoesNotExist:
        from django.http import JsonResponse
        return JsonResponse({'ok': False, 'erro': f'Chapa {chapa} não encontrada.'}, status=404)


@login_required
def api_salvar_localizacao(request, pk):
    """
    API JSON: atualiza a localização de um item patrimonial.
    Chamado pelo scanner após o usuário confirmar com "Salvar".
    POST body: {"localizacao_id": <int>}
    """
    from django.http import JsonResponse
    import json
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'erro': 'Método não permitido'}, status=405)
    try:
        data = json.loads(request.body)
        localizacao_id = data.get('localizacao_id')
        item = get_object_or_404(PatrimonioItem, pk=pk)

        loc_nome = ''
        if localizacao_id:
            loc = get_object_or_404(Localizacao, pk=localizacao_id)
            item.localizacao = loc
            loc_nome = loc.nome
        else:
            item.localizacao = None

        item.save(update_fields=['localizacao', 'atualizado_em'])

        LogAuditoria.registrar(
            acao=LogAuditoria.ACAO_EDITAR,
            tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
            descricao=f'Inventário scanner: localizou [{item.numero_chapa}] {item.nome} em "{loc_nome or "sem localização"}"',
            usuario=request.user,
            entidade_id=str(item.pk),
            entidade_nome=item.nome,
        )

        return JsonResponse({'ok': True, 'localizacao_nome': loc_nome})
    except Exception as e:
        return JsonResponse({'ok': False, 'erro': str(e)}, status=500)


@login_required
def buscar_por_chapa(request, numero_chapa):
    """
    Redireciona para o item com o número de chapa informado.
    Chamado pelo scanner após decodificar o QR Code.

    URL: /patrimonio/chapa/<numero>/
    """
    item = get_object_or_404(PatrimonioItem, numero_chapa=numero_chapa)
    return redirect('patrimonio_detalhe', pk=item.pk)


@login_required
def gerar_qrcode(request, pk):
    """
    Gera e retorna a imagem PNG do QR Code de um item patrimonial.
    O QR Code codifica: numero_chapa <TAB> data_aquisicao <TAB> descricao
    (mesmo formato de copiar 3 colunas de uma planilha Excel).

    URL: /patrimonio/<pk>/qrcode/
    Retorna: imagem PNG direta (para exibir em <img> ou fazer download)
    """
    import io
    import qrcode

    item = get_object_or_404(PatrimonioItem, pk=pk)

    data = item.data_aquisicao.strftime('%d/%m/%Y') if item.data_aquisicao else ''

    # Codifica os 3 campos separados por tabulação, igual ao copiar colunas do Excel
    conteudo = f'{item.numero_chapa}\t{data}\t{item.nome}'

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(conteudo)
    qr.make(fit=True)

    img = qr.make_image(fill_color='black', back_color='white')
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)

    return HttpResponse(buffer.read(), content_type='image/png')


@login_required
def exportar_csv(request):
    """
    Exporta todos os itens patrimoniais para um arquivo CSV.
    Formato de cada linha: ,numero_chapa,,data_aquisicao,,descricao,
    Responde com download direto do arquivo.

    URL: /patrimonio/exportar/csv/
    """
    import csv

    # Cria a resposta HTTP com tipo CSV e cabeçalho para download
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="patrimonio.csv"'

    writer = csv.writer(response, delimiter=',')

    # Busca todos os itens ordenados por número de chapa
    itens = PatrimonioItem.objects.all().order_by('numero_chapa')

    for item in itens:
        # Formata a data no padrão dd/mm/aaaa (ou vazio se não tiver)
        data = item.data_aquisicao.strftime('%d/%m/%Y') if item.data_aquisicao else ''
        descricao = item.descricao or ''

        # Formato: ,numero,,data,,descricao,
        writer.writerow(['', item.numero_chapa, '', data, '', descricao, ''])

    return response


@login_required
def pagina_qrcode(request, pk):
    """
    Página dedicada para visualizar e imprimir o QR Code de um item.
    Inclui informações do item e botão de impressão.

    URL: /patrimonio/<pk>/qrcode/imprimir/
    """
    item = get_object_or_404(PatrimonioItem, pk=pk)
    return render(request, 'patrimonio/qrcode_item.html', {
        'item': item,
        'pagina_ativa': 'patrimonio',
    })


@login_required
def etiqueta_print(request, pk):
    """
    Página standalone (sem base.html, sem Tailwind) para impressão da etiqueta térmica.
    Abre em nova janela e auto-imprime ao carregar.
    URL: /patrimonio/<pk>/etiqueta/print/
    """
    item = get_object_or_404(PatrimonioItem, pk=pk)
    return render(request, 'patrimonio/etiqueta_print.html', {'item': item})


# ==============================================================
# LEITOR QR CODE — CONFERÊNCIA AVULSA
# ==============================================================

@login_required
def leitor_qr_conferencia(request):
    """Página de captura de patrimônios via leitor QR (teclado HID)."""
    return render(request, 'patrimonio/leitor_qr_conferencia.html', {
        'pagina_ativa': 'leitor_qr',
    })


@login_required
def comparar_leitor_xls(request):
    """
    Recebe via POST:
      - numeros_json: JSON com lista de números de patrimônio lidos pelo leitor
      - arquivo_xls:  arquivo XLS/XLSX de referência para comparação

    Localiza a coluna de patrimônio no XLS (busca por 'chapa' ou 'patrimônio'
    no cabeçalho; se não encontrar, usa a primeira coluna).

    Gera um XLS de resultado com 3 abas:
      1. Encontrados    — lidos E presentes no XLS de referência
      2. Somente Lidos  — lidos MAS ausentes no XLS de referência
      3. Somente no XLS — no XLS de referência MAS não lidos
    """
    import json as _json
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import datetime

    if request.method != 'POST':
        return redirect('leitor_qr_conferencia')

    # ── 1. Ler lista do leitor ────────────────────────────────────────────────
    numeros_raw = request.POST.get('numeros_json', '[]')
    try:
        numeros_lidos = [str(n).strip() for n in _json.loads(numeros_raw) if str(n).strip()]
    except _json.JSONDecodeError:
        messages.error(request, 'Dados inválidos. Tente novamente.')
        return redirect('leitor_qr_conferencia')

    if not numeros_lidos:
        messages.warning(request, 'Nenhum patrimônio capturado.')
        return redirect('leitor_qr_conferencia')

    # ── 2. Ler XLS de referência ──────────────────────────────────────────────
    arquivo = request.FILES.get('arquivo_xls')
    if not arquivo:
        messages.error(request, 'Nenhum arquivo XLS enviado.')
        return redirect('leitor_qr_conferencia')

    try:
        wb_ref = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        ws_ref = wb_ref.active
    except Exception:
        messages.error(request, 'Não foi possível ler o arquivo XLS enviado.')
        return redirect('leitor_qr_conferencia')

    # Detectar coluna de patrimônio: busca 'chapa' ou 'patrimônio' no cabeçalho
    cabecalho = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws_ref.iter_rows())]
    col_pat = 0  # índice 0-based; default = primeira coluna
    for i, h in enumerate(cabecalho):
        if 'chapa' in h or 'patrim' in h:
            col_pat = i
            break

    # Ler todos os valores de patrimônio do XLS (a partir da linha 2)
    xls_rows = []   # lista de listas com todos os valores da linha
    for row in ws_ref.iter_rows(min_row=2, values_only=True):
        val = row[col_pat] if len(row) > col_pat else None
        if val is not None and str(val).strip():
            xls_rows.append({'num': str(val).strip(), 'row': list(row)})

    wb_ref.close()

    xls_set    = {r['num'] for r in xls_rows}
    lidos_set  = set(numeros_lidos)

    encontrados   = [r for r in xls_rows if r['num'] in lidos_set]
    so_xls        = [r for r in xls_rows if r['num'] not in lidos_set]
    so_lidos      = [n for n in numeros_lidos if n not in xls_set]

    # ── 3. Montar XLS de resultado ────────────────────────────────────────────
    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    centro = Alignment(horizontal='center', vertical='center')

    def estilo_cab(cel, cor_hex):
        cel.font      = Font(bold=True, color='FFFFFF')
        cel.fill      = PatternFill('solid', fgColor=cor_hex)
        cel.alignment = centro
        cel.border    = borda

    def estilo_dado(cel, cor_hex, alinhar_centro=False):
        cel.fill      = PatternFill('solid', fgColor=cor_hex)
        cel.alignment = Alignment(vertical='center',
                                  horizontal='center' if alinhar_centro else 'left')
        cel.border    = borda

    wb_out = openpyxl.Workbook()

    # ── Aba 1: Encontrados ────────────────────────────────────────────────────
    ws1 = wb_out.active
    ws1.title = 'Encontrados'
    ws1.append(['Nº Patrimônio'] + cabecalho)
    estilo_cab(ws1.cell(1, 1), '1D6A27')
    for i, h in enumerate(cabecalho, 2):
        estilo_cab(ws1.cell(1, i), '1D6A27')

    for r in encontrados:
        ws1.append([r['num']] + r['row'])
        for col in range(1, len(r['row']) + 2):
            estilo_dado(ws1.cell(ws1.max_row, col), 'C6EFCE', col == 1)

    ws1.column_dimensions['A'].width = 18

    # ── Aba 2: Somente Lidos ──────────────────────────────────────────────────
    ws2 = wb_out.create_sheet('Somente Lidos')
    ws2.append(['Nº Patrimônio', 'Observação'])
    estilo_cab(ws2.cell(1, 1), 'B45309')
    estilo_cab(ws2.cell(1, 2), 'B45309')

    for num in so_lidos:
        ws2.append([num, 'Lido mas ausente no XLS de referência'])
        estilo_dado(ws2.cell(ws2.max_row, 1), 'FFEB9C', True)
        estilo_dado(ws2.cell(ws2.max_row, 2), 'FFEB9C')

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 40

    # ── Aba 3: Somente no XLS ─────────────────────────────────────────────────
    ws3 = wb_out.create_sheet('Somente no XLS')
    ws3.append(['Nº Patrimônio'] + cabecalho)
    estilo_cab(ws3.cell(1, 1), '1E3A8A')
    for i in range(len(cabecalho)):
        estilo_cab(ws3.cell(1, i + 2), '1E3A8A')

    for r in so_xls:
        ws3.append([r['num']] + r['row'])
        for col in range(1, len(r['row']) + 2):
            estilo_dado(ws3.cell(ws3.max_row, col), 'DBEAFE', col == 1)

    ws3.column_dimensions['A'].width = 18

    # ── 4. Retornar arquivo ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)

    data_hora    = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    nome_arquivo = f'conferencia_patrimonio_{data_hora}.xlsx'

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response


# ==============================================================
# XLS DE REFERÊNCIA — AUTO-PREENCHIMENTO DO FORMULÁRIO
# ==============================================================

import json
from pathlib import Path
from django.conf import settings
from django.http import JsonResponse

# Caminho do arquivo JSON onde os dados do XLS ficam armazenados
XLS_REF_JSON = Path(settings.MEDIA_ROOT) / 'xls_referencia.json'


@login_required
def carregar_xls_referencia(request):
    """
    Recebe o upload do XLS de referência, extrai os dados e salva
    em um JSON local para consultas rápidas no formulário.

    GET  → exibe o formulário de upload
    POST → processa o arquivo e redireciona para criar novo item
    """
    import openpyxl

    # Informa quantos itens já estão carregados (se houver)
    try:
        total_carregado = XLSReferenciaItem.objects.count()
    except Exception:
        # Tabela ainda não existe (migration pendente) — roda migrate e tenta de novo
        from django.core.management import call_command
        call_command('migrate', '--noinput', verbosity=0)
        total_carregado = XLSReferenciaItem.objects.count()

    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        try:
            # data_only=True retorna os valores calculados das fórmulas
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active

            # Mapeamento de Estado 2 → status do sistema
            mapa_status = {
                'ótimo':    'ativo',
                'otimo':    'ativo',
                'bom':      'ativo',
                'regular':  'manutencao',
                'péssimo':  'baixado',
                'pessimo':  'baixado',
                'baixado':  'baixado',
            }

            objetos = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                chapa_raw = row[0]   # Patrim
                if not chapa_raw:
                    continue
                try:
                    chapa_int = int(float(chapa_raw))
                except (ValueError, TypeError):
                    continue

                # Data de aquisição — pode vir como datetime ou string
                data_raw = row[1]
                if hasattr(data_raw, 'strftime'):
                    data_iso = data_raw.strftime('%Y-%m-%d')
                elif data_raw:
                    try:
                        from datetime import datetime as dt
                        data_iso = dt.strptime(str(data_raw).strip(), '%d/%m/%Y').strftime('%Y-%m-%d')
                    except Exception:
                        data_iso = ''
                else:
                    data_iso = ''

                descricao = str(row[2]).strip() if len(row) > 2 and row[2] else ''  # Descricao
                local     = str(row[4]).strip() if len(row) > 4 and row[4] else ''  # Local 2
                estado2   = str(row[6]).strip().lower() if len(row) > 6 and row[6] else ''
                status    = mapa_status.get(estado2, '')

                objetos.append(XLSReferenciaItem(
                    numero_chapa=chapa_int,
                    nome=descricao,
                    data_aquisicao=data_iso,
                    local=local,
                    status=status,
                ))

            # Substitui todos os dados de uma vez
            XLSReferenciaItem.objects.all().delete()
            XLSReferenciaItem.objects.bulk_create(objetos, batch_size=500, ignore_conflicts=True)

            messages.success(request, f'XLS carregado com sucesso: {len(objetos)} itens disponíveis para auto-preenchimento.')
            return redirect('patrimonio_criar')

        except Exception as e:
            messages.error(request, f'Erro ao processar o arquivo: {e}')

    return render(request, 'patrimonio/carregar_xls_ref.html', {
        'pagina_ativa':     'patrimonio',
        'total_carregado':  total_carregado,
    })


@login_required
def buscar_dados_xls(request, chapa):
    """
    API JSON chamada pelo JavaScript do formulário.
    Retorna os dados do item pelo número de chapa, consultando o banco.

    URL: /patrimonio/xls-ref/<chapa>/
    """
    # Verifica se há dados carregados (para o banner)
    if chapa == 0 and not XLSReferenciaItem.objects.exists():
        return JsonResponse({'encontrado': False, 'erro': 'XLS não carregado'})

    try:
        item = XLSReferenciaItem.objects.filter(numero_chapa=chapa).first()
        if item:
            return JsonResponse({
                'encontrado': True,
                'nome':   item.nome,
                'data':   item.data_aquisicao,
                'local':  item.local,
                'status': item.status,
            })
        return JsonResponse({'encontrado': False})
    except Exception as e:
        return JsonResponse({'encontrado': False, 'erro': str(e)})


# ==============================================================
# MÓDULO DE CONFERÊNCIA DE INVENTÁRIO
# ==============================================================

def _carregar_xls_dados():
    """Retorna os dados do XLS de referência como dict {chapa: {...}}."""
    return {
        str(item.numero_chapa): {
            'nome':   item.nome,
            'data':   item.data_aquisicao,
            'local':  item.local,
            'status': item.status,
        }
        for item in XLSReferenciaItem.objects.all()
    }


@login_required
def conferencia_inicio(request):
    """
    Página inicial da conferência.
    - Admin vê todos os setores.
    - Conferente vê apenas os setores que tem permissão.
    """
    setores_permitidos = get_setores_do_conferente(request.user)
    is_conferente = setores_permitidos is not None  # None = admin (vê tudo)

    # Conferente sem nenhum setor atribuído
    if is_conferente and not setores_permitidos:
        messages.warning(request, 'Você ainda não tem nenhum setor de conferência atribuído. Solicite ao administrador.')
        return redirect('dashboard')

    dados = _carregar_xls_dados()
    if not dados and not is_conferente:
        messages.warning(request, 'Carregue o XLS de referência antes de iniciar a conferência.')
        return redirect('carregar_xls_referencia')

    # Agrupa itens por localização (vindos do XLS)
    locais = {}
    for chapa, item in dados.items():
        local = item.get('local', '').strip()
        if not local:
            continue
        if is_conferente and local not in setores_permitidos:
            continue
        locais[local] = locais.get(local, 0) + 1

    # Inclui localizações cadastradas no banco que não aparecem no XLS
    for loc in Localizacao.objects.all():
        if is_conferente and loc.nome not in setores_permitidos:
            continue
        if loc.nome not in locais:
            locais[loc.nome] = PatrimonioItem.objects.filter(localizacao=loc).count()

    # Para conferente: garante que setores atribuídos apareçam mesmo sem itens
    if is_conferente:
        for nome in setores_permitidos:
            if nome not in locais:
                locais[nome] = 0

    locais_lista = sorted(locais.items())  # Ordena A-Z

    return render(request, 'patrimonio/conferencia_inicio.html', {
        'locais':         locais_lista,
        'total_locais':   len(locais_lista),
        'total_itens':    sum(locais.values()),
        'is_conferente':  is_conferente,
        'pagina_ativa':   'conferencia',
    })


@login_required
def conferencia_cadastrar_item(request):
    """
    Cadastra com um clique um item somente_xls na conferência de sala.
    Usa os dados do XLSReferenciaItem (nome, data) para criar o PatrimonioItem
    sem exigir que o conferente preencha um formulário completo.
    POST: chapa=<int>  local_nome=<nome da sala>
    Disponível para admin e conferentes com permissão no setor.
    """
    from django.urls import reverse
    if request.method != 'POST':
        return redirect('conferencia_inicio')

    local_nome = request.POST.get('local_nome', '').strip()
    chapa_str  = request.POST.get('chapa', '').strip()

    if not local_nome or not chapa_str:
        return redirect('conferencia_inicio')

    if not pode_conferir_setor(request.user, local_nome):
        messages.error(request, 'Sem permissão para cadastrar itens neste setor.')
        return redirect('conferencia_inicio')

    try:
        chapa = int(chapa_str)
    except ValueError:
        messages.error(request, 'Número de chapa inválido.')
        return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")

    # Se o item já existe no banco, apenas transfere para cá (safety net)
    existente = PatrimonioItem.objects.filter(numero_chapa=chapa).first()
    if existente:
        loc, _ = Localizacao.objects.get_or_create(nome=local_nome)
        loc_anterior = existente.localizacao.nome if existente.localizacao else 'sem localização'
        existente.localizacao = loc
        existente.save(update_fields=['localizacao', 'atualizado_em'])
        xls_ref = XLSReferenciaItem.objects.filter(numero_chapa=chapa).first()
        if xls_ref:
            xls_ref.local = local_nome
            xls_ref.save(update_fields=['local'])
        prefixo = '[CONFERENTE] ' if not request.user.is_staff else ''
        LogAuditoria.registrar(
            acao=LogAuditoria.ACAO_TRANSFERIR,
            tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
            descricao=f'{prefixo}Conferência: transferiu [{existente.numero_chapa}] {existente.nome} de "{loc_anterior}" para "{local_nome}".',
            usuario=request.user,
            entidade_id=str(existente.pk),
            entidade_nome=existente.nome,
        )
        messages.success(request, f'Item [{chapa}] transferido para "{local_nome}".')
        return redirect(f"{reverse('conferencia_sala')}?local={local_nome}#chapa-{chapa}")

    # Item não existe no banco — cria usando dados do XLSReferenciaItem
    xls_ref = XLSReferenciaItem.objects.filter(numero_chapa=chapa).first()
    nome_item = xls_ref.nome if xls_ref and xls_ref.nome else f'Item {chapa}'
    data_aq = None
    if xls_ref and xls_ref.data_aquisicao:
        from datetime import date
        try:
            data_aq = date.fromisoformat(xls_ref.data_aquisicao)
        except ValueError:
            pass

    loc, _ = Localizacao.objects.get_or_create(nome=local_nome)
    novo = PatrimonioItem(
        numero_chapa=chapa,
        nome=nome_item,
        localizacao=loc,
        data_aquisicao=data_aq,
        criado_por=request.user,
        criado_por_nome=request.user.get_full_name() or request.user.username,
    )
    novo.save()

    # Garante que o XLS de referência registra este local
    if xls_ref:
        xls_ref.local = local_nome
        xls_ref.save(update_fields=['local'])
    else:
        XLSReferenciaItem.objects.create(
            numero_chapa=chapa,
            nome=nome_item,
            local=local_nome,
        )

    prefixo = '[CONFERENTE] ' if not request.user.is_staff else ''
    LogAuditoria.registrar(
        acao=LogAuditoria.ACAO_CRIAR,
        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
        descricao=f'{prefixo}Conferência: cadastrou [{chapa}] {nome_item} em "{local_nome}" via conferência de sala.',
        usuario=request.user,
        entidade_id=str(novo.pk),
        entidade_nome=novo.nome,
    )
    messages.success(request, f'Item [{chapa}] "{nome_item}" cadastrado em "{local_nome}".')
    return redirect(f"{reverse('conferencia_sala')}?local={local_nome}#chapa-{chapa}")


@login_required
def conferencia_transferir(request, pk):
    """
    Transfere um item patrimonial para a sala em conferência (traz para cá).
    Disponível para admin e conferentes com permissão no setor.
    POST: local_nome=<nome da sala>
    """
    from django.urls import reverse
    if request.method != 'POST':
        return redirect('conferencia_inicio')

    item = get_object_or_404(PatrimonioItem, pk=pk)
    local_nome = request.POST.get('local_nome', '').strip()

    if not local_nome:
        return redirect('conferencia_inicio')

    if not pode_conferir_setor(request.user, local_nome):
        messages.error(request, 'Sem permissão para transferir para este setor.')
        return redirect('conferencia_inicio')

    loc, _ = Localizacao.objects.get_or_create(nome=local_nome)
    loc_anterior = item.localizacao.nome if item.localizacao else 'sem localização'
    item.localizacao = loc
    if 'baixado' in local_nome.lower():
        item.status = PatrimonioItem.STATUS_BAIXADO
    item.save(update_fields=['localizacao', 'status', 'atualizado_em'])

    # Atualiza o XLS de referência para refletir a nova localização
    xls_item = XLSReferenciaItem.objects.filter(numero_chapa=item.numero_chapa).first()
    if xls_item:
        xls_item.local = local_nome
        xls_item.save(update_fields=['local'])
    else:
        XLSReferenciaItem.objects.create(
            numero_chapa=item.numero_chapa,
            nome=item.nome,
            data_aquisicao=item.data_aquisicao.strftime('%Y-%m-%d') if item.data_aquisicao else '',
            local=local_nome,
            status=item.status,
        )

    prefixo = '[CONFERENTE] ' if not request.user.is_staff else ''
    LogAuditoria.registrar(
        acao=LogAuditoria.ACAO_TRANSFERIR,
        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
        descricao=f'{prefixo}Conferência: transferiu [{item.numero_chapa}] {item.nome} de "{loc_anterior}" para "{local_nome}" (XLS atualizado).',
        usuario=request.user,
        entidade_id=str(item.pk),
        entidade_nome=item.nome,
    )
    messages.success(request, f'Item [{item.numero_chapa}] transferido para "{local_nome}" e XLS atualizado.')
    return redirect(f"{reverse('conferencia_sala')}?local={local_nome}#chapa-{item.numero_chapa}")


@login_required
def conferencia_transferir_lote(request):
    """
    Transfere um conjunto de itens (selecionados via checkbox) para uma localização destino.
    POST: local_nome=<sala origem>  destino_pk=<pk destino>  pks=<id> pks=<id> ...
    Disponível para admin e conferentes com permissão no setor de origem.
    """
    from django.urls import reverse

    if request.method != 'POST':
        return redirect('conferencia_inicio')

    local_nome  = request.POST.get('local_nome', '').strip()
    destino_pk  = request.POST.get('destino_pk', '').strip()
    pks         = request.POST.getlist('pks')

    if not local_nome or not destino_pk or not pks:
        messages.error(request, 'Dados incompletos para a transferência em lote.')
        return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")

    if not pode_conferir_setor(request.user, local_nome):
        messages.error(request, 'Sem permissão para transferir itens deste setor.')
        return redirect('conferencia_inicio')

    destino = get_object_or_404(Localizacao, pk=destino_pk)

    # Busca todos os itens de uma vez
    itens = list(PatrimonioItem.objects.filter(pk__in=pks).select_related('localizacao'))
    transferidos = len(itens)

    if transferidos == 0:
        messages.warning(request, 'Nenhum item encontrado para transferir.')
        return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")

    # Atualiza todos os PatrimonioItem em uma única query SQL
    update_kwargs = {'localizacao': destino}
    if 'baixado' in destino.nome.lower():
        update_kwargs['status'] = PatrimonioItem.STATUS_BAIXADO
    PatrimonioItem.objects.filter(pk__in=pks).update(**update_kwargs)

    # Atualiza o XLS em lote: busca todos os registros existentes de uma vez
    chapas = [item.numero_chapa for item in itens]
    xls_existentes = {
        x.numero_chapa: x
        for x in XLSReferenciaItem.objects.filter(numero_chapa__in=chapas)
    }

    xls_para_atualizar = []
    xls_para_criar = []
    for item in itens:
        if item.numero_chapa in xls_existentes:
            xls_item = xls_existentes[item.numero_chapa]
            xls_item.local = destino.nome
            xls_para_atualizar.append(xls_item)
        else:
            xls_para_criar.append(XLSReferenciaItem(
                numero_chapa=item.numero_chapa,
                nome=item.nome,
                data_aquisicao=item.data_aquisicao.strftime('%Y-%m-%d') if item.data_aquisicao else '',
                local=destino.nome,
                status=item.status,
            ))

    if xls_para_atualizar:
        XLSReferenciaItem.objects.bulk_update(xls_para_atualizar, ['local'])
    if xls_para_criar:
        XLSReferenciaItem.objects.bulk_create(xls_para_criar)

    prefixo = '[CONFERENTE] ' if not request.user.is_staff else ''
    LogAuditoria.registrar(
        acao=LogAuditoria.ACAO_TRANSFERIR,
        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
        descricao=(
            f'{prefixo}Conferência lote: transferiu {transferidos} item(ns) de "{local_nome}" '
            f'para "{destino.nome}" (XLS atualizado).'
        ),
        usuario=request.user,
        entidade_id='',
        entidade_nome=destino.nome,
    )

    messages.success(
        request,
        f'{transferidos} item(ns) transferido(s) para "{destino.nome}".'
    )
    return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")


@login_required
def conferencia_confirmar_xls(request, pk):
    """
    Para itens 'somente_db': confirma a presença do item nesta sala
    atualizando (ou criando) o registro no XLS de referência.
    Disponível para admin e conferentes com permissão no setor.
    POST: local_nome=<nome da sala>
    """
    from django.urls import reverse
    if request.method != 'POST':
        return redirect('conferencia_inicio')

    item = get_object_or_404(PatrimonioItem, pk=pk)
    local_nome = request.POST.get('local_nome', '').strip()
    if not local_nome:
        return redirect('conferencia_inicio')

    if not pode_conferir_setor(request.user, local_nome):
        messages.error(request, 'Sem permissão para confirmar itens neste setor.')
        return redirect('conferencia_inicio')

    xls_item, criado = XLSReferenciaItem.objects.get_or_create(
        numero_chapa=item.numero_chapa,
        defaults={
            'nome':           item.nome,
            'data_aquisicao': item.data_aquisicao.strftime('%Y-%m-%d') if item.data_aquisicao else '',
            'local':          local_nome,
            'status':         item.status,
        }
    )
    if not criado:
        xls_item.local = local_nome
        xls_item.save(update_fields=['local'])

    LogAuditoria.registrar(
        acao=LogAuditoria.ACAO_EDITAR,
        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
        descricao=f'Conferência: confirmou [{item.numero_chapa}] {item.nome} em "{local_nome}" no XLS de referência.',
        usuario=request.user,
        entidade_id=str(item.pk),
        entidade_nome=item.nome,
    )
    messages.success(request, f'Item [{item.numero_chapa}] confirmado em "{local_nome}" — agora aparece como Conferido.')
    return redirect(f"{reverse('conferencia_sala')}?local={local_nome}#chapa-{item.numero_chapa}")


@login_required
def conferencia_confirmar_xls_lote(request):
    """
    Confirma em lote todos os itens 'somente_db' da sala no XLS de referência.
    Disponível para admin e conferentes com permissão no setor.
    POST: local_nome=<nome da sala>
    """
    from django.urls import reverse
    if request.method != 'POST':
        return redirect('conferencia_inicio')

    local_nome = request.POST.get('local_nome', '').strip()
    if not local_nome:
        return redirect('conferencia_inicio')

    if not pode_conferir_setor(request.user, local_nome):
        messages.error(request, 'Sem permissão para confirmar itens neste setor.')
        return redirect('conferencia_inicio')

    # Chapas já presentes no XLS para esta sala
    chapas_xls = set(
        XLSReferenciaItem.objects.filter(local=local_nome)
        .values_list('numero_chapa', flat=True)
    )

    # Itens no banco nesta sala que NÃO estão no XLS
    itens_somente_db = PatrimonioItem.objects.filter(
        localizacao__nome=local_nome
    ).exclude(numero_chapa__in=chapas_xls).select_related('localizacao')

    total = 0
    for item in itens_somente_db:
        xls_item, criado = XLSReferenciaItem.objects.get_or_create(
            numero_chapa=item.numero_chapa,
            defaults={
                'nome':           item.nome,
                'data_aquisicao': item.data_aquisicao.strftime('%Y-%m-%d') if item.data_aquisicao else '',
                'local':          local_nome,
                'status':         item.status,
            }
        )
        if not criado and xls_item.local != local_nome:
            xls_item.local = local_nome
            xls_item.save(update_fields=['local'])
        total += 1

    if total:
        LogAuditoria.registrar(
            acao=LogAuditoria.ACAO_EDITAR,
            tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
            descricao=f'Conferência lote: confirmou {total} item(ns) de "{local_nome}" no XLS de referência.',
            usuario=request.user,
        )
        messages.success(request, f'{total} item(ns) confirmado(s) no XLS — agora aparecem como Conferido.')
    else:
        messages.info(request, 'Nenhum item pendente de confirmação no XLS.')

    return redirect(f"{reverse('conferencia_sala')}?local={local_nome}#fim")


@login_required
def conferencia_enviar_para_fora(request, pk):
    """
    Transfere um item DESTE setor para outro setor (sentido inverso do transferir para cá).
    Disponível para admin e conferentes com permissão no setor de origem.
    POST: local_nome=<sala atual>  destino_pk=<pk da localização destino>
    """
    from django.urls import reverse
    if request.method != 'POST':
        return redirect('conferencia_inicio')

    item = get_object_or_404(PatrimonioItem, pk=pk)
    local_nome = request.POST.get('local_nome', '').strip()
    destino_pk = request.POST.get('destino_pk', '').strip()

    if not local_nome or not destino_pk:
        messages.error(request, 'Dados incompletos para a transferência.')
        return redirect('conferencia_inicio')

    if not pode_conferir_setor(request.user, local_nome):
        messages.error(request, f'Sem permissão para mover itens do setor "{local_nome}".')
        return redirect('conferencia_inicio')

    destino = get_object_or_404(Localizacao, pk=destino_pk)
    loc_anterior = item.localizacao.nome if item.localizacao else 'sem localização'

    item.localizacao = destino
    if 'baixado' in destino.nome.lower():
        item.status = PatrimonioItem.STATUS_BAIXADO
    item.save(update_fields=['localizacao', 'status', 'atualizado_em'])

    # Atualiza o XLS de referência
    xls_item = XLSReferenciaItem.objects.filter(numero_chapa=item.numero_chapa).first()
    if xls_item:
        xls_item.local = destino.nome
        xls_item.save(update_fields=['local'])
    else:
        XLSReferenciaItem.objects.create(
            numero_chapa=item.numero_chapa,
            nome=item.nome,
            data_aquisicao=item.data_aquisicao.strftime('%Y-%m-%d') if item.data_aquisicao else '',
            local=destino.nome,
            status=item.status,
        )

    prefixo = '[CONFERENTE] ' if not request.user.is_staff else ''
    LogAuditoria.registrar(
        acao=LogAuditoria.ACAO_TRANSFERIR,
        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
        descricao=(
            f'{prefixo}Conferência: enviou [{item.numero_chapa}] {item.nome} '
            f'de "{loc_anterior}" para "{destino.nome}" (XLS atualizado).'
        ),
        usuario=request.user,
        entidade_id=str(item.pk),
        entidade_nome=item.nome,
    )
    messages.success(request, f'Item [{item.numero_chapa}] enviado para "{destino.nome}".')
    return redirect(f"{reverse('conferencia_sala')}?local={local_nome}#chapa-{item.numero_chapa}")


@login_required
def conferencia_importar_localizacoes(request):
    """
    Cria registros de Localizacao no banco para cada
    localização única encontrada no XLS de referência.
    Ignora as que já existem (pelo nome).
    """
    dados = _carregar_xls_dados()
    if not dados:
        messages.warning(request, 'Carregue o XLS de referência primeiro.')
        return redirect('carregar_xls_referencia')

    locais_unicos = {item.get('local', '').strip() for item in dados.values()}
    locais_unicos.discard('')

    criados = 0
    ignorados = 0
    for nome in sorted(locais_unicos):
        _, criado = Localizacao.objects.get_or_create(nome=nome)
        if criado:
            criados += 1
        else:
            ignorados += 1

    messages.success(
        request,
        f'{criados} localizações criadas, {ignorados} já existiam.'
    )
    return redirect('conferencia_inicio')


@login_required
def conferencia_sala(request):
    """
    Compara itens do XLS vs banco de dados para uma localização.
    Recebe o nome da sala via query param: ?local=Fisioterapia

    Classifica cada item em:
      - conferido:    chapa existe no XLS e no banco
      - somente_xls:  chapa no XLS mas não cadastrada no banco
      - somente_db:   chapa no banco (nesta localização) mas não no XLS

    Acesso: admin (qualquer sala) ou conferente (apenas seus setores autorizados).
    """
    local_nome = request.GET.get('local', '').strip()
    if not local_nome:
        return redirect('conferencia_inicio')

    if not pode_conferir_setor(request.user, local_nome):
        messages.error(request, f'Você não tem permissão para conferir o setor "{local_nome}".')
        return redirect('conferencia_inicio')

    dados = _carregar_xls_dados()

    # Itens do XLS para esta localização
    itens_xls = {
        int(chapa): item
        for chapa, item in dados.items()
        if item.get('local', '').strip() == local_nome
    }

    chapas_xls = set(itens_xls.keys())

    # Itens do banco com essas chapas
    itens_db_por_chapa = {
        item.numero_chapa: item
        for item in PatrimonioItem.objects.filter(numero_chapa__in=chapas_xls)
        .select_related('localizacao')
    }

    # Itens no banco com localização de mesmo nome, mas chapas não listadas no XLS
    itens_extras = PatrimonioItem.objects.filter(
        localizacao__nome=local_nome
    ).exclude(numero_chapa__in=chapas_xls).select_related('localizacao')

    # Monta lista de resultados
    resultados = []

    for chapa in sorted(itens_xls.keys()):
        xls = itens_xls[chapa]
        db  = itens_db_por_chapa.get(chapa)
        if db is None:
            status = 'somente_xls'
            localizacao_atual = ''
        else:
            loc_db_nome = db.localizacao.nome if db.localizacao else ''
            if loc_db_nome == local_nome:
                status = 'conferido'
            else:
                status = 'local_divergente'
            localizacao_atual = loc_db_nome
        item_status = db.status if db else ''
        resultados.append({
            'chapa':            chapa,
            'nome_xls':         xls.get('nome', ''),
            'nome_db':          db.nome if db else '',
            'data_xls':         xls.get('data', ''),
            'status':           status,
            'db_pk':            db.pk if db else None,
            'localizacao_atual': localizacao_atual,
            'item_status':      item_status,
            'is_baixado':       (
                item_status == 'baixado'
                or 'baixado' in localizacao_atual.lower()
                or 'baixado' in local_nome.lower()
            ),
        })

    for db in itens_extras:
        loc_atual = db.localizacao.nome if db.localizacao else ''
        item_status = db.status
        resultados.append({
            'chapa':            db.numero_chapa,
            'nome_xls':         '',
            'nome_db':          db.nome,
            'data_xls':         '',
            'status':           'somente_db',
            'db_pk':            db.pk,
            'localizacao_atual': loc_atual,
            'item_status':      item_status,
            'is_baixado':       (
                item_status == 'baixado'
                or 'baixado' in loc_atual.lower()
                or 'baixado' in local_nome.lower()
            ),
        })

    conferidos      = sum(1 for r in resultados if r['status'] == 'conferido')
    divergentes     = sum(1 for r in resultados if r['status'] == 'local_divergente')
    somente_xls     = sum(1 for r in resultados if r['status'] == 'somente_xls')
    somente_db      = sum(1 for r in resultados if r['status'] == 'somente_db')

    # Busca o log de conferência mais recente para cada item "conferido"
    pks_conferidos = [
        str(r['db_pk']) for r in resultados
        if r['status'] == 'conferido' and r['db_pk']
    ]
    logs_conferencia = {}
    if pks_conferidos:
        qs = (
            LogAuditoria.objects
            .filter(
                entidade_id__in=pks_conferidos,
                tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
                descricao__contains='Conferência:',
            )
            .order_by('entidade_id', '-criado_em')
        )
        for log in qs:
            if log.entidade_id not in logs_conferencia:
                logs_conferencia[log.entidade_id] = log

    for r in resultados:
        if r['status'] == 'conferido' and r['db_pk']:
            log = logs_conferencia.get(str(r['db_pk']))
            r['conferido_em']  = log.criado_em      if log else None
            r['conferido_por'] = log.usuario_nome   if log else None
        else:
            r['conferido_em']  = None
            r['conferido_por'] = None

    localizacoes = Localizacao.objects.exclude(nome=local_nome).order_by('nome')

    return render(request, 'patrimonio/conferencia_sala.html', {
        'local_nome':    local_nome,
        'resultados':    resultados,
        'conferidos':    conferidos,
        'divergentes':   divergentes,
        'somente_xls':   somente_xls,
        'somente_db':    somente_db,
        'total':         len(resultados),
        'localizacoes':  localizacoes,
        'is_conferente': not request.user.is_staff,
        'pagina_ativa':  'conferencia',
    })


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def conferencia_reset(request):
    """
    Transfere todos os itens da sala atual para uma localização escolhida,
    deixando a sala vazia no banco para que a conferência recomece do zero.
    O XLS de referência não é alterado — permanece como gabarito da sala.

    POST: local_nome=<sala origem>  destino_pk=<pk da localização destino>
    """
    from django.urls import reverse

    if request.method != 'POST':
        return redirect('conferencia_inicio')

    local_nome = request.POST.get('local_nome', '').strip()
    destino_pk = request.POST.get('destino_pk', '').strip()

    if not local_nome or not destino_pk:
        messages.error(request, 'Dados incompletos para o reset.')
        return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")

    destino = get_object_or_404(Localizacao, pk=destino_pk)

    itens = PatrimonioItem.objects.filter(localizacao__nome=local_nome).select_related('localizacao')
    total = itens.count()

    if total == 0:
        messages.info(request, f'Nenhum item em "{local_nome}" para transferir.')
        return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")

    reset_kwargs = {'localizacao': destino}
    if 'baixado' in destino.nome.lower():
        reset_kwargs['status'] = PatrimonioItem.STATUS_BAIXADO
    itens.update(**reset_kwargs)

    LogAuditoria.registrar(
        acao=LogAuditoria.ACAO_EDITAR,
        tipo_entidade=LogAuditoria.ENTIDADE_PATRIMONIO,
        descricao=(
            f'Conferência RESET: {total} item(ns) transferido(s) de "{local_nome}" '
            f'para "{destino.nome}". XLS de referência mantido.'
        ),
        usuario=request.user,
        entidade_id='',
        entidade_nome=local_nome,
    )

    messages.success(
        request,
        f'{total} item(ns) transferido(s) para "{destino.nome}". '
        f'A sala "{local_nome}" está zerada — conferência reiniciada.'
    )
    return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def conferencia_remover_xls(request):
    """
    Remove um XLSReferenciaItem criado por acidente (item 'somente_xls').
    POST: chapa=<numero_chapa>  local_nome=<nome da sala>
    """
    from django.urls import reverse

    if request.method != 'POST':
        return redirect('conferencia_inicio')

    local_nome = request.POST.get('local_nome', '').strip()
    chapa_raw  = request.POST.get('chapa', '').strip()

    try:
        chapa = int(chapa_raw)
    except (ValueError, TypeError):
        messages.error(request, 'Chapa inválida.')
        return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")

    deletado, _ = XLSReferenciaItem.objects.filter(
        numero_chapa=chapa, local=local_nome
    ).delete()

    if deletado:
        messages.success(request, f'Item {chapa} removido do XLS de referência.')
    else:
        messages.warning(request, f'Item {chapa} não encontrado no XLS para "{local_nome}".')

    return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")


@login_required
def conferencia_exportar(request):
    """
    Gera o Excel de conferência para uma sala.
    Recebe o nome via ?local=Fisioterapia
    Gera: Fisioterapia_conferido.xlsx
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    local_nome = request.GET.get('local', '').strip()
    if not local_nome:
        return redirect('conferencia_inicio')

    dados = _carregar_xls_dados()

    itens_xls = {
        int(chapa): item
        for chapa, item in dados.items()
        if item.get('local', '').strip() == local_nome
    }
    chapas_xls = set(itens_xls.keys())

    itens_db_por_chapa = {
        item.numero_chapa: item
        for item in PatrimonioItem.objects.filter(numero_chapa__in=chapas_xls)
    }
    itens_extras = PatrimonioItem.objects.filter(
        localizacao__nome=local_nome
    ).exclude(numero_chapa__in=chapas_xls)

    # Monta linhas
    linhas = []
    for chapa in sorted(itens_xls.keys()):
        xls = itens_xls[chapa]
        db  = itens_db_por_chapa.get(chapa)
        linhas.append({
            'chapa':    chapa,
            'nome':     xls.get('nome', ''),
            'data':     xls.get('data', ''),
            'local':    local_nome,
            'status':   'Conferido' if db else 'Somente no XLS',
        })
    for db in itens_extras:
        linhas.append({
            'chapa':    db.numero_chapa,
            'nome':     db.nome,
            'data':     db.data_aquisicao.strftime('%d/%m/%Y') if db.data_aquisicao else '',
            'local':    local_nome,
            'status':   'Somente no Banco',
        })

    # Cria o workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = local_nome[:31]  # Excel limita nome da aba a 31 chars

    # Cores
    verde  = PatternFill('solid', fgColor='C6EFCE')
    amarelo = PatternFill('solid', fgColor='FFEB9C')
    azul   = PatternFill('solid', fgColor='BDD7EE')
    cinza  = PatternFill('solid', fgColor='D9D9D9')

    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # Cabeçalho
    cabecalho = ['Chapa', 'Nome do Item', 'Data Aquisição', 'Localização', 'Status']
    for col, titulo in enumerate(cabecalho, 1):
        cel = ws.cell(row=1, column=col, value=titulo)
        cel.font      = Font(bold=True, color='FFFFFF')
        cel.fill      = PatternFill('solid', fgColor='1E40AF')
        cel.alignment = Alignment(horizontal='center')
        cel.border    = borda

    # Dados
    cor_status = {
        'Conferido':        verde,
        'Somente no XLS':   amarelo,
        'Somente no Banco': azul,
    }
    for row_idx, linha in enumerate(linhas, 2):
        valores = [linha['chapa'], linha['nome'], linha['data'], linha['local'], linha['status']]
        fill = cor_status.get(linha['status'], cinza)
        for col, val in enumerate(valores, 1):
            cel = ws.cell(row=row_idx, column=col, value=val)
            cel.fill   = fill
            cel.border = borda
            if col == 1:
                cel.alignment = Alignment(horizontal='center')

    # Ajusta largura das colunas
    larguras = [10, 50, 18, 30, 20]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    # Aba de resumo
    ws2 = wb.create_sheet('Resumo')
    conferidos  = sum(1 for l in linhas if l['status'] == 'Conferido')
    s_xls       = sum(1 for l in linhas if l['status'] == 'Somente no XLS')
    s_db        = sum(1 for l in linhas if l['status'] == 'Somente no Banco')
    resumo = [
        ('Localização', local_nome),
        ('Total de itens', len(linhas)),
        ('Conferidos (XLS + Banco)', conferidos),
        ('Somente no XLS (não cadastrados)', s_xls),
        ('Somente no Banco (não no XLS)', s_db),
    ]
    for r, (label, valor) in enumerate(resumo, 1):
        ws2.cell(r, 1, label).font = Font(bold=True)
        ws2.cell(r, 2, valor)
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 30

    # Gera resposta HTTP com o arquivo
    nome_arquivo = f"{local_nome}_conferido.xlsx".replace('/', '-').replace('\\', '-')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    wb.save(response)
    return response


# ==============================================================
# EXPORTAR QR CODES EM LOTE — ITENS CONFERIDOS DE UMA SALA
# ==============================================================

@login_required
def conferencia_exportar_qrcodes(request):
    """
    Gera um arquivo ZIP com os QR Codes PNG de todos os itens
    'conferidos' (presentes no XLS e no banco nesta localização).
    Cada PNG é nomeado com o número de chapa: <numero_chapa>.png
    Mesmo conteúdo do QR Code gerado por gerar_qrcode():
      numero_chapa TAB data_aquisicao TAB nome

    URL: /conferencia/exportar-qrcodes/?local=<nome>
    """
    import io
    import zipfile
    import qrcode

    local_nome = request.GET.get('local', '').strip()
    if not local_nome:
        return redirect('conferencia_inicio')

    dados = _carregar_xls_dados()

    # Chapas do XLS para esta sala
    chapas_xls = {
        int(chapa)
        for chapa, item in dados.items()
        if item.get('local', '').strip() == local_nome
    }

    # Itens no banco que estão nesta sala E estão no XLS (= conferidos)
    itens_conferidos = PatrimonioItem.objects.filter(
        numero_chapa__in=chapas_xls,
        localizacao__nome=local_nome,
    ).order_by('numero_chapa')

    if not itens_conferidos.exists():
        messages.warning(request, 'Nenhum item conferido encontrado para esta sala.')
        from django.urls import reverse
        return redirect(f"{reverse('conferencia_sala')}?local={local_nome}")

    # Monta o ZIP em memória
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for item in itens_conferidos:
            # Data do banco; se vazia, busca no XLS de referência
            if item.data_aquisicao:
                data = item.data_aquisicao.strftime('%d/%m/%Y')
            else:
                data_iso = dados.get(str(item.numero_chapa), {}).get('data', '')
                if data_iso:
                    try:
                        from datetime import datetime as _dt
                        data = _dt.strptime(data_iso, '%Y-%m-%d').strftime('%d/%m/%Y')
                    except ValueError:
                        data = data_iso  # usa como está se o formato for diferente
                else:
                    data = ''
            conteudo = f'{item.numero_chapa}\t{data}\t{item.nome}'

            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10,
                border=2,
            )
            qr.add_data(conteudo)
            qr.make(fit=True)
            img = qr.make_image(fill_color='black', back_color='white')

            png_buffer = io.BytesIO()
            img.save(png_buffer, format='PNG')
            zf.writestr(f'{item.numero_chapa}.png', png_buffer.getvalue())

    zip_buffer.seek(0)

    nome_arquivo = f"QRCodes_{local_nome}.zip".replace('/', '-').replace('\\', '-').replace(' ', '_')
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response


# ==============================================================
# COMPARAR DOIS XLS
# ==============================================================

@login_required
def comparar_xls(request):
    """
    Recebe dois arquivos XLS/XLSX e compara os itens pelo número de chapa.
    Mostra: itens em A apenas, itens em B apenas, itens em ambos com divergência.
    """
    import openpyxl

    resultado = None

    if request.method == 'POST':
        arquivo_a = request.FILES.get('arquivo_a')
        arquivo_b = request.FILES.get('arquivo_b')
        nome_a    = request.POST.get('nome_a', 'Arquivo A')
        nome_b    = request.POST.get('nome_b', 'Arquivo B')

        if arquivo_a and arquivo_b:
            def ler_xls(arquivo):
                """Lê um XLS e retorna dict {chapa: {'nome', 'data', 'local'}}"""
                wb = openpyxl.load_workbook(arquivo, data_only=True)
                ws = wb.active
                dados = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    chapa_raw = row[0]
                    if not chapa_raw:
                        continue
                    try:
                        chapa = str(int(float(chapa_raw)))
                    except (ValueError, TypeError):
                        continue
                    nome = str(row[9]).strip() if row[9] else (str(row[2]).strip() if row[2] else '')
                    data_raw = row[1]
                    if hasattr(data_raw, 'strftime'):
                        data = data_raw.strftime('%d/%m/%Y')
                    else:
                        data = str(data_raw).strip() if data_raw else ''
                    local = str(row[4]).strip() if row[4] else ''
                    dados[chapa] = {'nome': nome, 'data': data, 'local': local}
                return dados

            try:
                dados_a = ler_xls(arquivo_a)
                dados_b = ler_xls(arquivo_b)

                chapas_a = set(dados_a.keys())
                chapas_b = set(dados_b.keys())

                so_em_a    = []
                so_em_b    = []
                divergentes = []
                iguais     = 0

                for chapa in sorted(chapas_a | chapas_b, key=lambda x: int(x)):
                    em_a = chapa in chapas_a
                    em_b = chapa in chapas_b

                    if em_a and not em_b:
                        so_em_a.append({'chapa': chapa, **dados_a[chapa]})
                    elif em_b and not em_a:
                        so_em_b.append({'chapa': chapa, **dados_b[chapa]})
                    else:
                        # Em ambos: verifica divergências
                        diffs = []
                        for campo in ('nome', 'data', 'local'):
                            va = dados_a[chapa].get(campo, '')
                            vb = dados_b[chapa].get(campo, '')
                            if va != vb:
                                diffs.append({'campo': campo, 'valor_a': va, 'valor_b': vb})
                        if diffs:
                            divergentes.append({
                                'chapa': chapa,
                                'nome':  dados_a[chapa].get('nome', ''),
                                'diffs': diffs,
                            })
                        else:
                            iguais += 1

                resultado = {
                    'nome_a':      nome_a,
                    'nome_b':      nome_b,
                    'total_a':     len(dados_a),
                    'total_b':     len(dados_b),
                    'so_em_a':     so_em_a,
                    'so_em_b':     so_em_b,
                    'divergentes': divergentes,
                    'iguais':      iguais,
                }
            except Exception as e:
                messages.error(request, f'Erro ao processar os arquivos: {e}')

    return render(request, 'patrimonio/comparar_xls.html', {
        'resultado':    resultado,
        'pagina_ativa': 'patrimonio',
    })


# ----------------------------------------------------------
# ETIQUETAS COLORIDAS
# ----------------------------------------------------------
@login_required
def etiquetas_view(request):
    """Página para gerar e imprimir etiquetas coloridas 5×3 cm."""
    return render(request, 'patrimonio/etiquetas.html', {
        'pagina_ativa': 'etiquetas',
    })


# ==============================================================
# MANUTENÇÃO
# ==============================================================

@login_required
def manutencao_lista(request):
    """Lista todas as solicitações de manutenção."""
    filtro = request.GET.get('filtro', 'pendente')
    qs = ManutencaoSolicitacao.objects.all()
    if filtro == 'pendente':
        qs = qs.filter(status='pendente')
    elif filtro == 'concluido':
        qs = qs.filter(status='concluido')
    # filtro == 'todos' → sem filtro

    pendentes  = ManutencaoSolicitacao.objects.filter(status='pendente').count()
    concluidos = ManutencaoSolicitacao.objects.filter(status='concluido').count()

    return render(request, 'patrimonio/manutencao_lista.html', {
        'solicitacoes': qs,
        'filtro':       filtro,
        'pendentes':    pendentes,
        'concluidos':   concluidos,
        'pagina_ativa': 'manutencao',
    })


@login_required
def manutencao_registrar(request):
    """Recebe AJAX POST e salva a solicitação; também marca o item como 'manutencao'."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)

    chapa    = request.POST.get('chapa', '').strip()
    nome     = request.POST.get('nome_item', '').strip()
    sala     = request.POST.get('sala', '').strip()
    descricao = request.POST.get('descricao', '').strip()

    if not chapa or not sala or not descricao:
        return JsonResponse({'ok': False, 'erro': 'Dados incompletos'}, status=400)

    try:
        chapa_int = int(chapa)
    except ValueError:
        return JsonResponse({'ok': False, 'erro': 'Chapa inválida'}, status=400)

    ManutencaoSolicitacao.objects.create(
        numero_chapa=chapa_int,
        nome_item=nome,
        sala=sala,
        descricao=descricao,
    )

    # Atualiza status do item para 'manutencao'
    PatrimonioItem.objects.filter(numero_chapa=chapa_int).update(
        status=PatrimonioItem.STATUS_MANUTENCAO
    )

    return JsonResponse({'ok': True})


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def manutencao_concluir(request, pk):
    """Marca a solicitação como concluída e reverte o item para 'ativo'."""
    from django.utils import timezone
    sol = get_object_or_404(ManutencaoSolicitacao, pk=pk)
    sol.status       = ManutencaoSolicitacao.STATUS_CONCLUIDO
    sol.concluido_em = timezone.now()
    sol.save()

    # Só volta para ativo se não houver outras solicitações pendentes para o mesmo item
    outras_pendentes = ManutencaoSolicitacao.objects.filter(
        numero_chapa=sol.numero_chapa,
        status=ManutencaoSolicitacao.STATUS_PENDENTE,
    ).exists()
    if not outras_pendentes:
        PatrimonioItem.objects.filter(numero_chapa=sol.numero_chapa).update(
            status=PatrimonioItem.STATUS_ATIVO
        )

    messages.success(request, f'Manutenção do item {sol.numero_chapa} marcada como concluída.')
    return redirect(f'{request.META.get("HTTP_REFERER", "/manutencao/")}')


@login_required
@user_passes_test(is_admin, login_url='dashboard')
def manutencao_apagar(request, pk):
    """Apaga definitivamente uma solicitação."""
    sol = get_object_or_404(ManutencaoSolicitacao, pk=pk)
    chapa = sol.numero_chapa
    sol.delete()
    messages.success(request, f'Solicitação do item {chapa} removida.')
    return redirect(f'{request.META.get("HTTP_REFERER", "/manutencao/")}')
