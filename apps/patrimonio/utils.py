"""
============================================================
UTILS - Funções auxiliares do sistema
============================================================
Aqui ficam funções que não pertencem a nenhum model ou view
específico, mas são usadas por vários lugares do sistema.

Funções disponíveis:
  - processar_arquivo()   → lê Excel/CSV e retorna lista de dicts
  - gerar_pdf_relatorio() → gera PDF de relatório de patrimônio
  - formatar_valor()      → formata número como moeda brasileira
"""

import io
import csv
import datetime
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse


# ============================================================
# PROCESSAMENTO DE ARQUIVO (Excel / CSV)
# ============================================================

def processar_arquivo(arquivo):
    """
    Lê um arquivo Excel (.xlsx, .xls) ou CSV e retorna uma lista
    de dicionários, onde cada dicionário representa uma linha/item.

    Colunas esperadas no arquivo (nomes flexíveis, não case-sensitive):
      - numero_chapa / chapa / patrimônio / nº
      - nome / descrição / item
      - categoria / tipo
      - localizacao / local / setor
      - responsavel / responsável
      - data_aquisicao / data / aquisição
      - valor / valor_aquisicao / preço
      - status
      - descricao / observacao / obs

    Retorna:
      - lista de dicts com os itens encontrados
      - Ou lança ValueError com mensagem de erro legível
    """
    nome_arquivo = arquivo.name.lower()

    if nome_arquivo.endswith('.csv'):
        return _processar_csv(arquivo)
    elif nome_arquivo.endswith('.xlsx') or nome_arquivo.endswith('.xls'):
        return _processar_excel(arquivo)
    else:
        raise ValueError('Formato de arquivo não suportado. Use .xlsx, .xls ou .csv')


def _processar_excel(arquivo):
    """Processa arquivo Excel usando openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError('Biblioteca openpyxl não instalada. Execute: pip install openpyxl')

    try:
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb.active  # Usa a primeira planilha
    except Exception as e:
        raise ValueError(f'Erro ao abrir o arquivo Excel: {e}')

    # Lê o cabeçalho (primeira linha)
    cabecalho = []
    for cell in ws[1]:
        valor = str(cell.value or '').strip().lower()
        cabecalho.append(valor)

    if not any(cabecalho):
        raise ValueError('O arquivo está vazio ou sem cabeçalho na primeira linha.')

    itens = []
    # Lê as linhas de dados (a partir da segunda linha)
    for num_linha, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Pula linhas completamente vazias
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        linha_dict = {}
        for i, valor in enumerate(row):
            if i < len(cabecalho):
                linha_dict[cabecalho[i]] = valor

        item = _mapear_linha(linha_dict, num_linha)
        if item:
            itens.append(item)

    if not itens:
        raise ValueError('Nenhum item encontrado no arquivo. Verifique se há dados abaixo do cabeçalho.')

    return itens


def _processar_csv(arquivo):
    """Processa arquivo CSV."""
    try:
        # Tenta decodificar como UTF-8, depois como latin-1 (comum em arquivos brasileiros)
        try:
            conteudo = arquivo.read().decode('utf-8-sig')  # utf-8-sig remove o BOM se presente
        except UnicodeDecodeError:
            arquivo.seek(0)
            conteudo = arquivo.read().decode('latin-1')

        reader = csv.DictReader(io.StringIO(conteudo))

        # Normaliza os nomes das colunas para minúsculas
        itens = []
        for num_linha, row in enumerate(reader, start=2):
            linha_normalizada = {k.strip().lower(): v for k, v in row.items()}
            item = _mapear_linha(linha_normalizada, num_linha)
            if item:
                itens.append(item)

    except Exception as e:
        raise ValueError(f'Erro ao processar CSV: {e}')

    if not itens:
        raise ValueError('Nenhum item encontrado no arquivo CSV.')

    return itens


def _mapear_linha(linha, num_linha):
    """
    Mapeia uma linha do arquivo para um dicionário padronizado.
    Aceita variações nos nomes das colunas (flexível).

    Retorna um dict com campos padronizados, ou None se a linha for inválida.
    """

    def pegar(chaves):
        """Tenta pegar o valor de um conjunto de nomes de coluna possíveis."""
        for chave in chaves:
            if chave in linha and linha[chave] not in (None, ''):
                return str(linha[chave]).strip()
        return ''

    # Tenta encontrar o número da chapa
    numero_chapa_str = pegar([
        'numero_chapa', 'num_chapa', 'chapa', 'patrimônio', 'patrimonio',
        'patrim', 'nº patrim', 'n° patrim', 'nº', 'n°', 'numero', 'número',
        'id', 'código', 'codigo'
    ])

    # O nome do item é obrigatório
    nome = pegar(['nome', 'descrição', 'descricao', 'item', 'bem', 'objeto'])
    if not nome:
        return None  # Pula linhas sem nome

    # Converte número da chapa
    numero_chapa = None
    if numero_chapa_str:
        try:
            numero_chapa = int(float(numero_chapa_str))
        except (ValueError, TypeError):
            pass  # Se não conseguir converter, deixa None (será gerado automaticamente)

    # Converte valor monetário
    valor_str = pegar(['valor', 'valor_aquisicao', 'preço', 'preco', 'custo'])
    valor = None
    if valor_str:
        # Remove R$, pontos de milhar e troca vírgula por ponto
        valor_limpo = valor_str.replace('R$', '').replace('.', '').replace(',', '.').strip()
        try:
            valor = Decimal(valor_limpo)
        except InvalidOperation:
            valor = None

    # Converte data
    data_str = pegar(['data_aquisicao', 'data', 'aquisição', 'aquisicao', 'dt_aquisicao'])
    data_aquisicao = _converter_data(data_str)

    # Status
    status_str = pegar(['status', 'situação', 'situacao']).lower()
    status_mapa = {
        'ativo': 'ativo', 'active': 'ativo', '1': 'ativo',
        'manutenção': 'manutencao', 'manutencao': 'manutencao', 'manutenção': 'manutencao',
        'baixado': 'baixado', 'inativo': 'baixado', 'inactive': 'baixado', '0': 'baixado',
    }
    status = status_mapa.get(status_str, 'ativo')

    # Localização: se vier como float (ex: 1.0), converte para inteiro string ("1")
    loc_raw = pegar(['localizacao', 'localização', 'local', 'setor', 'departamento'])
    try:
        loc_nome = str(int(float(loc_raw))) if loc_raw else ''
    except (ValueError, TypeError):
        loc_nome = loc_raw

    return {
        'numero_chapa': numero_chapa,
        'nome': nome,
        'categoria': pegar(['categoria', 'tipo', 'type', 'group', 'grupo']),
        'localizacao_nome': loc_nome,
        'responsavel': pegar(['responsavel', 'responsável', 'resp']),
        'data_aquisicao': data_aquisicao,
        'valor': valor,
        'status': status,
        'descricao': pegar(['descricao', 'descrição', 'observacao', 'observação', 'obs']),
        '_linha': num_linha,  # Número da linha no arquivo (para mensagens de erro)
    }


def _converter_data(data_str):
    """
    Tenta converter uma string de data em objeto date.
    Suporta vários formatos comuns em arquivos brasileiros.
    """
    if not data_str or str(data_str).strip() == '':
        return None

    data_str = str(data_str).strip()

    # Trata número serial do Excel (dias desde 01/01/1900)
    try:
        serial = float(data_str)
        if 1000 < serial < 100000:  # Parece um número serial do Excel
            origem = datetime.date(1899, 12, 30)
            return origem + datetime.timedelta(days=int(serial))
    except (ValueError, TypeError):
        pass

    # Lista de formatos para tentar
    formatos = [
        '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',   # Formatos BR: 31/12/2024
        '%Y-%m-%d', '%Y/%m/%d',                 # Formato ISO: 2024-12-31
        '%d/%m/%y', '%d-%m-%y',                 # Com ano curto: 31/12/24
    ]
    for fmt in formatos:
        try:
            return datetime.datetime.strptime(data_str, fmt).date()
        except ValueError:
            continue

    return None  # Não conseguiu converter


# ============================================================
# GERAÇÃO DE PDF
# ============================================================

def gerar_pdf_relatorio(itens, titulo='Relatório de Patrimônio', filtros=None):
    """
    Gera um PDF com a lista de itens patrimoniais.

    Parâmetros:
      - itens   → queryset ou lista de PatrimonioItem
      - titulo  → título do relatório
      - filtros → dict com filtros aplicados (ex: {'Status': 'Ativo'})

    Retorna:
      - HttpResponse com o PDF pronto para download
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    # Cria buffer em memória (não salva em disco)
    buffer = io.BytesIO()

    # Configura o documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),   # Paisagem para caber mais colunas
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    # Estilos de texto
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'Titulo',
        parent=estilos['Title'],
        fontSize=16,
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    estilo_subtitulo = ParagraphStyle(
        'Subtitulo',
        parent=estilos['Normal'],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
        alignment=TA_CENTER,
    )
    estilo_celula = ParagraphStyle(
        'Celula',
        parent=estilos['Normal'],
        fontSize=7,
        wordWrap='CJK',
    )

    elementos = []

    # --- Título ---
    elementos.append(Paragraph(titulo, estilo_titulo))
    data_hora = datetime.datetime.now().strftime('%d/%m/%Y às %H:%M')
    elementos.append(Paragraph(f'Gerado em {data_hora}', estilo_subtitulo))

    # --- Filtros aplicados ---
    if filtros:
        texto_filtros = ' | '.join([f'{k}: {v}' for k, v in filtros.items() if v])
        if texto_filtros:
            elementos.append(Paragraph(f'Filtros: {texto_filtros}', estilo_subtitulo))

    elementos.append(Spacer(1, 0.3 * cm))

    # --- Estatísticas resumo ---
    total = len(itens) if hasattr(itens, '__len__') else itens.count()
    ativos = sum(1 for i in itens if i.status == 'ativo')
    manutencao = sum(1 for i in itens if i.status == 'manutencao')
    baixados = sum(1 for i in itens if i.status == 'baixado')
    valor_total = sum(
        (i.valor or 0) for i in itens if i.valor
    )

    resumo_data = [
        ['Total de Itens', 'Ativos', 'Em Manutenção', 'Baixados', 'Valor Total'],
        [
            str(total),
            str(ativos),
            str(manutencao),
            str(baixados),
            f'R$ {valor_total:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
        ]
    ]
    tabela_resumo = Table(resumo_data, colWidths=[4 * cm] * 5)
    tabela_resumo.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWHEIGHT', (0, 0), (-1, -1), 18),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#eff6ff')),
    ]))
    elementos.append(tabela_resumo)
    elementos.append(Spacer(1, 0.5 * cm))

    # --- Tabela principal de itens ---
    cabecalho_tabela = [
        'Chapa', 'Nome', 'Categoria', 'Localização', 'Responsável',
        'Data Aquis.', 'Valor (R$)', 'Status'
    ]

    STATUS_LABELS = {
        'ativo': 'Ativo',
        'manutencao': 'Manutenção',
        'baixado': 'Baixado',
    }

    dados_tabela = [cabecalho_tabela]
    for item in itens:
        valor_fmt = ''
        if item.valor:
            v = float(item.valor)
            valor_fmt = f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

        data_fmt = ''
        if item.data_aquisicao:
            data_fmt = item.data_aquisicao.strftime('%d/%m/%Y')

        dados_tabela.append([
            str(item.numero_chapa),
            Paragraph(item.nome[:60], estilo_celula),
            item.categoria[:25] if item.categoria else '',
            str(item.localizacao) if item.localizacao else '',
            item.responsavel[:20] if item.responsavel else '',
            data_fmt,
            valor_fmt,
            STATUS_LABELS.get(item.status, item.status),
        ])

    # Larguras das colunas (total deve caber em A4 paisagem ≈ 27cm úteis)
    col_widths = [1.8*cm, 6*cm, 3.5*cm, 3.5*cm, 3*cm, 2.2*cm, 2.5*cm, 2.2*cm]
    tabela = Table(dados_tabela, colWidths=col_widths, repeatRows=1)

    # Estilo da tabela
    estilo_tabela = TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Dados
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWHEIGHT', (0, 0), (-1, -1), 16),
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1e3a8a')),
        # Linhas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Chapa centralizada
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),   # Valor alinhado à direita
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),  # Status centralizado
    ])
    tabela.setStyle(estilo_tabela)
    elementos.append(tabela)

    # --- Rodapé ---
    elementos.append(Spacer(1, 0.5 * cm))
    elementos.append(Paragraph(
        'Sistema de Gestão Patrimonial - Relatório gerado automaticamente',
        ParagraphStyle('rodape', parent=estilos['Normal'], fontSize=7,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    # Gera o PDF
    doc.build(elementos)

    # Prepara o nome do arquivo para download
    buffer.seek(0)
    nome_arquivo = f'relatorio_patrimonio_{datetime.date.today():%Y%m%d}.pdf'

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response


# ============================================================
# FUNÇÕES AUXILIARES DE FORMATAÇÃO
# ============================================================

def formatar_valor(valor):
    """
    Formata um Decimal/float como moeda brasileira.
    Ex: 1234.56 → 'R$ 1.234,56'
    """
    if valor is None:
        return ''
    try:
        v = float(valor)
        return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return ''
